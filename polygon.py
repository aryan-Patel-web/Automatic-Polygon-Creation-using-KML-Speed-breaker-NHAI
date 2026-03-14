"""
polygon.py  v7  —  Speed Breaker CAP PTBM Polygon Engine
IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI Road Safety Automation

FEATURES:
  ● Per-marker configuration: each marker can have its own:
      - num_lanes   (1=town-village, 2=city-town, 4=city-city)
      - road_width_m
      - separator_width_m
      - heading_deg
  ● 3-tier heading detection: per-marker → global → OSM → PCA → neighbour
  ● Lane-aware strip placement with symmetric separator gap
  ● Parallel strip guarantee: all strips at one marker share exact heading
  ● Curvature detection via bearing-delta analysis
  ● KML export with per-lane colour styles + curve pin highlighting
  ● Excel BOQ: 5 sheets — Annexure1, Heading Assignments, Marker Details,
                            Strip Coordinates, Project Spec

GEOMETRY (per MoRTH/NHAI CAP PTBM specification):
  strip LENGTH  = lane_width  (across road, perpendicular to heading)
  strip WIDTH   = strip_mm    (along road, thin dimension 10/15mm)
  strips stacked ALONG road, centred at marker
  all strips at same marker → identical heading → perfectly parallel
"""

import math
import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict, Any
import urllib.request
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

R_EARTH = 6_371_000.0

LANE_PRESETS = {
    1: {"label": "1-Lane (Town→Village)", "road_width": 3.5,  "separator": 0.0, "has_sep": False},
    2: {"label": "2-Lane (City→Town)",    "road_width": 7.0,  "separator": 0.5, "has_sep": True},
    4: {"label": "4-Lane (City→City)",    "road_width": 14.0, "separator": 2.0, "has_sep": True},
    6: {"label": "6-Lane (Highway)",      "road_width": 21.0, "separator": 3.0, "has_sep": True},
}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class MarkerInfo:
    """One KML point placemark."""
    index: int
    name: str
    lat: float
    lon: float
    description: str = ""
    placement_code: str = ""


@dataclass
class MarkerOverride:
    """
    Per-marker configuration that overrides the global PolygonSpec.
    Any field set to None means "use global spec value".
    """
    heading_deg: Optional[float]       = None   # road heading 0-179°
    num_lanes: Optional[int]           = None   # 1, 2, 4, 6
    road_width_m: Optional[float]      = None   # total road width
    separator_width_m: Optional[float] = None   # centre divider width
    has_separator: Optional[bool]      = None
    lane_gap_m: Optional[float]        = None   # extra gap between lane groups
    strip_length_m: Optional[float]    = None   # strip length override


@dataclass
class PolygonSpec:
    """
    Global defaults for all markers.
    Per-marker overrides in MarkerOverride take precedence.
    """
    # Strip dimensions
    strip_width_mm: float           = 15.0
    num_strips: int                 = 6
    gap_between_strips_m: float     = 0.10
    strip_length_override_m: float  = -1.0    # -1 = auto (= lane_width)

    # Road defaults (used when no per-marker override)
    num_lanes: int                  = 2
    road_width_m: float             = 7.0
    separator_width_m: float        = 0.5
    has_separator: bool             = True
    lane_gap_m: float               = -1.0    # -1 = auto

    # Heading
    heading_override: float         = -1.0    # global; -1 = auto-detect

    def effective_lane_gap(self, lane_w: float, sep_w: float) -> float:
        """Total clear zone between lane groups (separator + clearance)."""
        if self.lane_gap_m > 0:
            return self.lane_gap_m
        return sep_w + max(0.3, lane_w * 0.10)

    def resolve(self, override: Optional[MarkerOverride]) -> Dict[str, Any]:
        """
        Merge global spec with per-marker override.
        Returns a dict with resolved values for this specific marker.
        """
        if override is None:
            override = MarkerOverride()
        nl       = override.num_lanes       if override.num_lanes       is not None else self.num_lanes
        rw       = override.road_width_m    if override.road_width_m    is not None else self.road_width_m
        sw       = override.separator_width_m if override.separator_width_m is not None else self.separator_width_m
        hs       = override.has_separator   if override.has_separator   is not None else self.has_separator
        lg       = override.lane_gap_m      if override.lane_gap_m      is not None else self.lane_gap_m
        sl       = override.strip_length_m  if override.strip_length_m  is not None else self.strip_length_override_m
        drv      = rw - (sw if hs and nl > 1 else 0.0)
        lw       = drv / max(1, nl)
        eff_gap  = (sw + max(0.3, lw * 0.10)) if (hs and nl > 1 and lg <= 0) else max(lg, 0.0)
        return {
            "num_lanes": nl,
            "road_width_m": rw,
            "separator_width_m": sw,
            "has_separator": hs,
            "lane_gap_m": eff_gap,
            "lane_width_m": lw,
            "strip_length_m": sl if sl > 0 else lw,
        }


@dataclass
class GeneratedPolygon:
    marker: MarkerInfo
    coordinates: List[Tuple[float, float]]           # bounding hull (lon, lat)
    heading_deg: float
    road_curvature: str
    strip_polygons: List[List[Tuple[float, float]]]  # per-strip closed rings
    lane_assignments: List[int]
    heading_source: str = "auto"
    num_lanes_used: int = 2
    spec: PolygonSpec = field(default_factory=PolygonSpec)
    override: Optional[MarkerOverride] = None


# ═══════════════════════════════════════════════════════════════════════════════
# KML PARSING
# ═══════════════════════════════════════════════════════════════════════════════

def parse_kml_markers(kml_path: str) -> List[MarkerInfo]:
    """Parse all Point Placemarks from a KML file."""
    tree = ET.parse(kml_path)
    root = tree.getroot()
    markers, idx = [], 1
    for pm in root.iter():
        if pm.tag.split('}')[-1] != 'Placemark':
            continue
        ne = _ch(pm, 'name')
        de = _ch(pm, 'description')
        name = (ne.text or '').strip() if ne is not None else f"Marker_{idx}"
        desc = (de.text or '').strip() if de is not None else ""
        coord = _extract_point(pm)
        if coord is None:
            continue
        lon, lat = coord
        code = name if any(k in name.upper()
                           for k in ['CAP','PTBM','MM','FM','SIGN','GO SLOW']) else ""
        markers.append(MarkerInfo(idx, name, lat, lon, desc, code))
        idx += 1
    return markers


def _ch(el, tag):
    for c in el:
        if c.tag.split('}')[-1] == tag:
            return c
    return None


def _extract_point(pm) -> Optional[Tuple[float, float]]:
    for el in pm.iter():
        if el.tag.split('}')[-1] == 'coordinates' and el.text:
            p = el.text.strip().split(',')
            if len(p) >= 2:
                try:
                    return float(p[0]), float(p[1])
                except ValueError:
                    pass
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# GEODESIC MATHEMATICS
# ═══════════════════════════════════════════════════════════════════════════════

def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in metres."""
    φ1, φ2 = math.radians(lat1), math.radians(lat2)
    a = (math.sin(math.radians(lat2 - lat1) / 2) ** 2 +
         math.cos(φ1) * math.cos(φ2) *
         math.sin(math.radians(lon2 - lon1) / 2) ** 2)
    return 2 * R_EARTH * math.asin(math.sqrt(max(0.0, min(1.0, a))))


def forward_bearing(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Forward azimuth [0, 360)."""
    φ1, φ2 = math.radians(lat1), math.radians(lat2)
    dλ = math.radians(lon2 - lon1)
    x = math.sin(dλ) * math.cos(φ2)
    y = math.cos(φ1) * math.sin(φ2) - math.sin(φ1) * math.cos(φ2) * math.cos(dλ)
    return (math.degrees(math.atan2(x, y)) + 360) % 360


def offset_point(lat: float, lon: float, dist_m: float, hdg_deg: float) -> Tuple[float, float]:
    """Move (lat, lon) by dist_m in direction hdg_deg. Returns (lat, lon)."""
    if dist_m == 0:
        return lat, lon
    d = dist_m / R_EARTH
    h = math.radians(hdg_deg)
    φ1 = math.radians(lat)
    λ1 = math.radians(lon)
    φ2 = math.asin(math.sin(φ1) * math.cos(d) +
                   math.cos(φ1) * math.sin(d) * math.cos(h))
    λ2 = λ1 + math.atan2(math.sin(h) * math.sin(d) * math.cos(φ1),
                          math.cos(d) - math.sin(φ1) * math.sin(φ2))
    return math.degrees(φ2), math.degrees(λ2)


def normalise_heading(h: float) -> float:
    """Reduce heading to [0, 180) — road is bidirectional."""
    h = h % 360
    return h - 180 if h >= 180 else h


# ═══════════════════════════════════════════════════════════════════════════════
# HEADING DETECTION  (3-tier fallback)
# ═══════════════════════════════════════════════════════════════════════════════

def _osm_road_heading(lat: float, lon: float, radius_m: int = 40) -> Optional[float]:
    """
    Query OpenStreetMap Overpass API for nearest highway segment.
    Returns road heading [0, 180) or None on failure/timeout.
    """
    query = (f'[out:json][timeout:8];'
             f'way(around:{radius_m},{lat:.6f},{lon:.6f})[highway];out geom 8;')
    try:
        data = urllib.parse.urlencode({'data': query}).encode()
        req  = urllib.request.Request('https://overpass-api.de/api/interpreter', data)
        req.add_header('User-Agent', 'GIS-BOQ-Tool-v7')
        with urllib.request.urlopen(req, timeout=8) as r:
            ways = json.loads(r.read().decode()).get('elements', [])
        best_h, best_d = None, 1e9
        for way in ways:
            geom = way.get('geometry', [])
            for i in range(len(geom) - 1):
                n1, n2 = geom[i], geom[i + 1]
                ml = (n1['lat'] + n2['lat']) / 2
                mo = (n1['lon'] + n2['lon']) / 2
                d  = haversine_distance(lat, lon, ml, mo)
                if d < best_d:
                    best_d = d
                    b = forward_bearing(n1['lat'], n1['lon'], n2['lat'], n2['lon'])
                    best_h = normalise_heading(b)
        return best_h
    except Exception:
        return None


_osm_cache: Dict[Tuple[float, float], Optional[float]] = {}

def _osm_cached(lat: float, lon: float) -> Optional[float]:
    key = (round(lat, 4), round(lon, 4))
    if key not in _osm_cache:
        _osm_cache[key] = _osm_road_heading(lat, lon)
    return _osm_cache[key]


def pca_heading(markers: List[MarkerInfo]) -> Optional[float]:
    """
    Principal Component Analysis on all marker positions.
    Returns dominant road direction [0, 180) if spread > 10m, else None.
    """
    if len(markers) < 2:
        return None
    lats = [m.lat for m in markers]
    lons = [m.lon for m in markers]
    clat = sum(lats) / len(lats)
    clon = sum(lons) / len(lons)
    slat = 111_320.0
    slon = 111_320.0 * math.cos(math.radians(clat))
    ys = [(la - clat) * slat for la in lats]
    xs = [(lo - clon) * slon for lo in lons]
    spread = math.sqrt((max(xs) - min(xs)) ** 2 + (max(ys) - min(ys)) ** 2)
    if spread < 10.0:
        return None
    n  = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    cxx = sum((x - mx) ** 2 for x in xs) / n
    cxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / n
    cyy = sum((y - my) ** 2 for y in ys) / n
    tr   = cxx + cyy
    det  = cxx * cyy - cxy ** 2
    disc = math.sqrt(max(0.0, (tr / 2) ** 2 - det))
    lam  = tr / 2 + disc
    if abs(cxy) > 1e-12:
        vx, vy = lam - cyy, cxy
    elif cxx >= cyy:
        vx, vy = 1.0, 0.0
    else:
        vx, vy = 0.0, 1.0
    return normalise_heading((math.degrees(math.atan2(vx, vy)) + 360) % 360)


def _neighbour_bearing(markers: List[MarkerInfo], idx: int) -> float:
    """
    Weighted circular mean of road bearings from ±3 neighbours.
    Normalised to [0, 180) before averaging to prevent reversal bug.
    """
    n = len(markers)
    cur = markers[idx]
    ws = wc = tw = 0.0
    for off in range(-3, 4):
        if off == 0:
            continue
        j = idx + off
        if not 0 <= j < n:
            continue
        nb = markers[j]
        b  = forward_bearing(nb.lat, nb.lon, cur.lat, cur.lon) if off < 0 \
             else forward_bearing(cur.lat, cur.lon, nb.lat, nb.lon)
        nm = normalise_heading(b)
        w  = 1.0 / abs(off)
        ws += w * math.sin(math.radians(nm))
        wc += w * math.cos(math.radians(nm))
        tw += w
    if tw == 0:
        return 0.0
    return (math.degrees(math.atan2(ws / tw, wc / tw)) + 360) % 360


def resolve_heading(
    markers: List[MarkerInfo],
    idx: int,
    spec: PolygonSpec,
    pca_h: Optional[float],
    use_osm: bool,
    per_marker_headings: Dict[int, float],
) -> Tuple[float, str]:
    """
    Heading priority:
      1. per_marker_headings[marker_index]  → "per-marker"
      2. spec.heading_override              → "global"
      3. OSM Overpass API                   → "osm"
      4. PCA of all markers                 → "pca"
      5. Neighbour bearing average          → "neighbour"
    Returns (heading [0,180), source_label).
    """
    mk_idx = markers[idx].index
    if mk_idx in per_marker_headings:
        return normalise_heading(per_marker_headings[mk_idx]), "per-marker"
    if 0 <= spec.heading_override < 180:
        return float(spec.heading_override), "global"
    if use_osm:
        h = _osm_cached(markers[idx].lat, markers[idx].lon)
        if h is not None:
            return h, "osm"
    if pca_h is not None:
        return pca_h, "pca"
    return _neighbour_bearing(markers, idx), "neighbour"


def detect_curvature(markers: List[MarkerInfo], idx: int) -> str:
    """Classify road curvature at marker from bearing-delta of neighbours."""
    n = len(markers)
    if idx == 0 or idx >= n - 1:
        return "straight"
    b_in  = forward_bearing(markers[idx-1].lat, markers[idx-1].lon,
                             markers[idx].lat,   markers[idx].lon)
    b_out = forward_bearing(markers[idx].lat,    markers[idx].lon,
                             markers[idx+1].lat,  markers[idx+1].lon)
    delta = abs(normalise_heading(b_in) - normalise_heading(b_out))
    delta = min(delta, 180 - delta)
    if delta < 5:   return "straight"
    if delta < 20:  return "slight_curve"
    return "sharp_curve"


# ═══════════════════════════════════════════════════════════════════════════════
# STRIP RECTANGLE BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
#
#  Road  ──────────────────────────────────────────────► heading H
#
#  ←lane_w→  ←─ sep ─→  ←lane_w→
#  ┌────────┐            ┌────────┐
#  │════════│            │════════│  strip 1 & (n/2+1)
#  │════════│            │════════│  strip 2 & (n/2+2)
#  │════════│            │════════│  strip 3 & (n/2+3)
#  └────────┘            └────────┘
#
#  Each ═══ rectangle:
#    long side  = strip_length_m  (across road, ⊥ to H)
#    short side = strip_width_mm  (along road, ∥ to H)
#
#  pa, pb  = signed perpendicular extents from road centre
#            (+ve = left of heading,  −ve = right of heading)
# ═══════════════════════════════════════════════════════════════════════════════

def make_strip(
    mk_lat:   float,
    mk_lon:   float,
    hdg:      float,    # road heading [0, 360)
    along_m:  float,    # strip centre offset along road from marker
    sw_m:     float,    # strip width (thin dim, e.g. 0.015)
    pa:       float,    # near perpendicular edge (signed metres)
    pb:       float,    # far  perpendicular edge (signed metres)
) -> List[Tuple[float, float]]:
    """
    Build one strip rectangle. Returns closed ring [(lon, lat), ...] — 5 pts.
    """
    h_fwd  = hdg
    h_bwd  = (hdg + 180) % 360
    h_left = (hdg - 90 + 360) % 360
    h_rgt  = (hdg + 90) % 360
    half   = sw_m / 2.0

    sc_lat, sc_lon = offset_point(mk_lat, mk_lon, along_m, h_fwd)
    fe_lat, fe_lon = offset_point(sc_lat, sc_lon, half, h_fwd)
    be_lat, be_lon = offset_point(sc_lat, sc_lon, half, h_bwd)

    def gp(lat, lon, d):
        return offset_point(lat, lon,  d, h_left) if d >= 0 \
               else offset_point(lat, lon, -d, h_rgt)

    fn = gp(fe_lat, fe_lon, pa)
    ff = gp(fe_lat, fe_lon, pb)
    bf = gp(be_lat, be_lon, pb)
    bn = gp(be_lat, be_lon, pa)
    ring = [(fn[1], fn[0]), (ff[1], ff[0]), (bf[1], bf[0]), (bn[1], bn[0])]
    ring.append(ring[0])
    return ring


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN POLYGON GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_polygon_for_marker(
    markers:              List[MarkerInfo],
    idx:                  int,
    spec:                 PolygonSpec,
    pca_h:                Optional[float],
    use_osm:              bool,
    per_marker_headings:  Dict[int, float],
    per_marker_overrides: Dict[int, MarkerOverride],
) -> GeneratedPolygon:
    """
    Build all speed breaker strips at one marker.

    Uses per-marker overrides (lanes, road width, separator, heading)
    if provided, otherwise falls back to global PolygonSpec values.

    Cross-section (2-lane example, 6 strips = 3 per lane):
      [L1-S1][L1-S2][L1-S3]  ← SEP →  [L2-S4][L2-S5][L2-S6]
      All strips have SAME heading → guaranteed perfectly parallel.
    """
    cur  = markers[idx]
    mk_i = cur.index

    # Resolve heading
    hdg, src = resolve_heading(markers, idx, spec, pca_h, use_osm, per_marker_headings)
    curv = detect_curvature(markers, idx)

    # Resolve per-marker road config
    ov   = per_marker_overrides.get(mk_i)
    cfg  = spec.resolve(ov)

    sw_m      = spec.strip_width_mm / 1000.0
    gap_m     = spec.gap_between_strips_m
    nl        = max(1, cfg["num_lanes"])
    lane_w    = cfg["lane_width_m"]
    strip_len = cfg["strip_length_m"]
    half_gap  = cfg["lane_gap_m"] / 2.0 if (cfg["has_separator"] and nl > 1) else 0.0
    half_road = cfg["road_width_m"] / 2.0

    # Distribute strips evenly across lanes
    base = spec.num_strips // nl
    rem  = spec.num_strips % nl
    spl  = [base + (1 if i < rem else 0) for i in range(nl)]

    # Along-road positions: use per-lane count (NOT global count)
    max_spl      = max(spl)
    total_along  = max_spl * sw_m + (max_spl - 1) * gap_m
    start_along  = -total_along / 2.0

    strips: List[List[Tuple[float, float]]] = []
    lanes:  List[int]                        = []

    for lane_idx in range(nl):
        n_in = spl[lane_idx]
        if n_in == 0:
            continue

        # ── Perpendicular extents of this lane ────────────────────────────
        if nl == 1:
            pa, pb = -strip_len / 2.0, strip_len / 2.0

        elif lane_idx % 2 == 0:
            # Left-side lanes (+perp direction)
            tier        = lane_idx // 2
            inner_edge  = half_gap + tier * lane_w
            lane_centre = inner_edge + lane_w / 2.0
            pa = lane_centre - strip_len / 2.0
            pb = lane_centre + strip_len / 2.0

        else:
            # Right-side lanes (−perp direction)
            tier        = lane_idx // 2
            inner_edge  = half_gap + tier * lane_w
            lane_centre = -(inner_edge + lane_w / 2.0)
            pa = lane_centre + strip_len / 2.0
            pb = lane_centre - strip_len / 2.0

        # Build strips with LOCAL s_idx so all lanes align along road
        for s_idx in range(n_in):
            along = start_along + s_idx * (sw_m + gap_m) + sw_m / 2.0
            strips.append(make_strip(cur.lat, cur.lon, hdg, along, sw_m, pa, pb))
            lanes.append(lane_idx + 1)

    all_pts = [pt for s in strips for pt in s]
    bnd     = convex_hull(all_pts) if len(all_pts) >= 3 else all_pts

    return GeneratedPolygon(
        marker=cur,
        coordinates=bnd,
        heading_deg=hdg,
        road_curvature=curv,
        strip_polygons=strips,
        lane_assignments=lanes,
        heading_source=src,
        num_lanes_used=nl,
        spec=spec,
        override=ov,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# CONVEX HULL  (Graham scan)
# ═══════════════════════════════════════════════════════════════════════════════

def convex_hull(pts: List[Tuple[float, float]]) -> List[Tuple[float, float]]:
    pts = sorted(set(pts))
    if len(pts) <= 2:
        return pts + ([pts[0]] if pts else [])
    def cross(O, A, B):
        return (A[0]-O[0])*(B[1]-O[1]) - (A[1]-O[1])*(B[0]-O[0])
    lo, up = [], []
    for p in pts:
        while len(lo) >= 2 and cross(lo[-2], lo[-1], p) <= 0: lo.pop()
        lo.append(p)
    for p in reversed(pts):
        while len(up) >= 2 and cross(up[-2], up[-1], p) <= 0: up.pop()
        up.append(p)
    h = lo[:-1] + up[:-1]
    return h + [h[0]] if h else h


# ═══════════════════════════════════════════════════════════════════════════════
# KML EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

_KML_STYLES = """
  <Style id="sL1">
    <LineStyle><color>ff00d7ff</color><width>1</width></LineStyle>
    <PolyStyle><color>d000d7ff</color></PolyStyle>
  </Style>
  <Style id="sL2">
    <LineStyle><color>ff0088ff</color><width>1</width></LineStyle>
    <PolyStyle><color>d00088ff</color></PolyStyle>
  </Style>
  <Style id="sL3">
    <LineStyle><color>ff00ff88</color><width>1</width></LineStyle>
    <PolyStyle><color>d000ff88</color></PolyStyle>
  </Style>
  <Style id="sL4">
    <LineStyle><color>ff8800ff</color><width>1</width></LineStyle>
    <PolyStyle><color>d08800ff</color></PolyStyle>
  </Style>
  <Style id="bnd">
    <LineStyle><color>ff0000ff</color><width>2</width></LineStyle>
    <PolyStyle><color>110000ff</color></PolyStyle>
  </Style>
  <Style id="pin">
    <IconStyle>
      <color>ff00d7ff</color><scale>1.1</scale>
      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/ylw-circle.png</href></Icon>
    </IconStyle>
    <LabelStyle><color>ffffffff</color><scale>0.85</scale></LabelStyle>
  </Style>
  <Style id="pinCurve">
    <IconStyle>
      <color>ff0000ff</color><scale>1.2</scale>
      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-circle.png</href></Icon>
    </IconStyle>
    <LabelStyle><color>ffffffff</color><scale>0.85</scale></LabelStyle>
  </Style>
  <Style id="pinOverride">
    <IconStyle>
      <color>ff0088ff</color><scale>1.15</scale>
      <Icon><href>http://maps.google.com/mapfiles/kml/paddle/orange-circle.png</href></Icon>
    </IconStyle>
    <LabelStyle><color>ffffffff</color><scale>0.85</scale></LabelStyle>
  </Style>
"""


def _cs(coords, alt=0):
    return " ".join(f"{lo},{la},{alt}" for lo, la in coords)


def export_kml(
    markers:              List[MarkerInfo],
    polygons:             List[GeneratedPolygon],
    out_path:             str,
    spec:                 PolygonSpec,
    per_marker_headings:  Dict[int, float] = None,
    per_marker_overrides: Dict[int, MarkerOverride] = None,
):
    per_marker_headings  = per_marker_headings  or {}
    per_marker_overrides = per_marker_overrides or {}
    lane_styles = ["#sL1", "#sL2", "#sL3", "#sL4"]

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '<Document>',
        '  <n>CAP PTBM Speed Breaker Polygons v7</n>',
        '  <description>GIS BOQ Tool — IIIT Nagpur | PWD/NHAI</description>',
        _KML_STYLES,
    ]

    for pg in polygons:
        mk  = pg.marker
        ov  = per_marker_overrides.get(mk.index)
        has_hdg_ov  = mk.index in per_marker_headings
        has_lane_ov = ov is not None and ov.num_lanes is not None
        pin_style   = ("#pinCurve"   if has_hdg_ov  else
                       "#pinOverride" if has_lane_ov else "#pin")

        nl  = pg.num_lanes_used
        cfg = spec.resolve(ov)
        lw  = cfg["lane_width_m"]
        sep = f"{cfg['separator_width_m']:.1f}m" if cfg['has_separator'] and nl > 1 else "none"

        info = (
            f"<b>{mk.placement_code or mk.name}</b><br/>"
            f"{'🔶 PER-MARKER OVERRIDE<br/>' if (has_hdg_ov or has_lane_ov) else ''}"
            f"Heading: <b>{pg.heading_deg:.1f}°</b> [{pg.heading_source}]<br/>"
            f"Lanes: <b>{nl}</b> ({LANE_PRESETS.get(nl, {}).get('label', '')})<br/>"
            f"Road: {pg.road_curvature.replace('_', ' ').title()}<br/>"
            f"Strip: {spec.strip_width_mm}mm × {lw:.2f}m | Sep: {sep}<br/>"
            f"Total strips: {len(pg.strip_polygons)} | "
            f"Coords: {mk.lat:.6f}, {mk.lon:.6f}"
        )

        lines += [
            f'  <Folder>',
            f'    <n>{mk.name}</n>',
            '    <Placemark>',
            f'      <n>{mk.name}</n>',
            f'      <styleUrl>{pin_style}</styleUrl>',
            f'      <description><![CDATA[{info}]]></description>',
            f'      <Point><coordinates>{mk.lon},{mk.lat},0</coordinates></Point>',
            '    </Placemark>',
        ]

        if len(pg.coordinates) >= 3:
            lines += [
                '    <Placemark>',
                f'      <n>{mk.name} – Bounding</n>',
                '      <styleUrl>#bnd</styleUrl>',
                '      <Polygon><tessellate>1</tessellate>',
                '        <outerBoundaryIs><LinearRing>',
                f'          <coordinates>{_cs(pg.coordinates)}</coordinates>',
                '        </LinearRing></outerBoundaryIs></Polygon>',
                '    </Placemark>',
            ]

        for i, (strip, ln) in enumerate(
                zip(pg.strip_polygons, pg.lane_assignments), 1):
            st = lane_styles[(ln - 1) % len(lane_styles)]
            lines += [
                '    <Placemark>',
                f'      <n>{mk.name} L{ln}S{i}</n>',
                f'      <styleUrl>{st}</styleUrl>',
                f'      <description><![CDATA['
                f'Strip {i} | Lane {ln} | {spec.strip_width_mm}mm × {lw:.2f}m | '
                f'Hdg:{pg.heading_deg:.1f}°[{pg.heading_source}]'
                f']]></description>',
                '      <Polygon><tessellate>1</tessellate>',
                '        <outerBoundaryIs><LinearRing>',
                f'          <coordinates>{_cs(strip)}</coordinates>',
                '        </LinearRing></outerBoundaryIs></Polygon>',
                '    </Placemark>',
            ]

        lines.append('  </Folder>')

    lines += ['</Document>', '</kml>']
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL BOQ EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_excel(
    polygons:             List[GeneratedPolygon],
    out_path:             str,
    spec:                 PolygonSpec,
    per_marker_headings:  Dict[int, float] = None,
    per_marker_overrides: Dict[int, MarkerOverride] = None,
):
    per_marker_headings  = per_marker_headings  or {}
    per_marker_overrides = per_marker_overrides or {}

    wb   = openpyxl.Workbook()
    HF   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    YF   = PatternFill("solid", start_color="FFD700", end_color="FFD700")
    OF   = PatternFill("solid", start_color="FFA500", end_color="FFA500")
    AF   = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
    WF   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    BLF  = PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8")
    GRF  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    T    = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'),  bottom=Side(style='thin'))
    CC   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LC   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def hc(ws, r, c, v, fill=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = fill or HF
        cell.alignment = CC
        cell.border = T
        return cell

    def dc(ws, r, c, v, fill=None, al=None, bold=False, sz=9):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(size=sz, bold=bold)
        cell.fill = fill or WF
        cell.alignment = al or CC
        cell.border = T
        return cell

    # ── Sheet 1: BOQ Annexure 1 ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "BOQ Sheet 1"

    ws.merge_cells("A1:K1")
    ws['A1'].value = "ANNEXURE 1 - ESTIMATE"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = HF; ws['A1'].alignment = CC

    ws.merge_cells("A2:K2")
    ws['A2'].value = "Speed Breaker CAP PTBM Road Marking — Bill of Quantities"
    ws['A2'].font = Font(bold=True, size=11); ws['A2'].alignment = CC

    ws.merge_cells("A3:K3")
    ws['A3'].value = "GIS BOQ Tool v7 | IIIT Nagpur | Dr. Neha Kasture | PWD/NHAI"
    ws['A3'].font = Font(italic=True, size=9, color="555555"); ws['A3'].alignment = CC
    ws.row_dimensions[4].height = 4

    for ci, h in enumerate([
        "S.No.", "Description", "Code", "Nos.",
        "Length\n(m)", "Breadth\n(m)", "Area\n(Sqm)",
        "Total Qty\n(Sqm)", "Rate\n(Rs./Sqm)", "Amount\n(Rs.)",
        "Notes / Override"
    ], 1):
        hc(ws, 5, ci, h)
    ws.row_dimensions[5].height = 40

    ws.merge_cells("A6:K6")
    ws['A6'].value = (
        f"  Strip:{spec.strip_width_mm}mm | Strips:{spec.num_strips} | "
        f"Default: {spec.num_lanes}-lane/{spec.road_width_m}m | "
        f"Sep:{spec.separator_width_m}m | Gap:{spec.gap_between_strips_m}m"
    )
    ws['A6'].font = Font(bold=True, size=9, color="1F3864")
    ws['A6'].fill = BLF; ws['A6'].alignment = LC; ws['A6'].border = T

    swm = spec.strip_width_mm / 1000.0
    for ri, pg in enumerate(polygons, 1):
        mk  = pg.marker
        er  = ri + 6
        ov  = per_marker_overrides.get(mk.index)
        cfg = spec.resolve(ov)
        lw  = cfg["lane_width_m"]
        is_ov = (mk.index in per_marker_headings or ov is not None)
        fill  = OF if is_ov else (AF if ri % 2 == 0 else WF)

        a1 = swm * lw
        ta = a1 * spec.num_strips
        code = mk.placement_code or f"CAP PTBM {int(spec.strip_width_mm)}MM X {spec.num_strips}"
        desc = (f"Supply and Application of Single Rib Pattern Cold Applied Plastic (CAP) "
                f"Parabolic Transverse Bar Rumble Marking (PTBM) — "
                f"{int(spec.strip_width_mm)}mm Thk.")
        lane_lbl = LANE_PRESETS.get(pg.num_lanes_used, {}).get("label", f"{pg.num_lanes_used}-Lane")
        notes = (
            f"{pg.road_curvature.replace('_', ' ').title()} | "
            f"Hdg:{pg.heading_deg:.1f}°[{pg.heading_source}] | "
            f"{lane_lbl} | {mk.lat:.5f},{mk.lon:.5f}"
        )

        dc(ws, er, 1,  ri,     fill)
        dc(ws, er, 2,  desc,   fill, al=LC)
        dc(ws, er, 3,  code,   fill)
        dc(ws, er, 4,  1,      fill)
        dc(ws, er, 5,  round(cfg["road_width_m"], 2), fill)
        dc(ws, er, 6,  round(swm, 4),  fill)
        dc(ws, er, 7,  round(a1, 5),   fill)
        dc(ws, er, 8,  round(ta, 4),   fill)
        dc(ws, er, 9,  "",             fill)
        amt = ws.cell(row=er, column=10, value=f'=IF(I{er}="","",H{er}*I{er})')
        amt.fill = fill; amt.border = T; amt.alignment = CC; amt.font = Font(size=9)
        dc(ws, er, 11, notes, fill, al=LC)

    tr = len(polygons) + 7
    ws.merge_cells(f"A{tr}:G{tr}")
    for ci in range(1, 12):
        c = ws.cell(row=tr, column=ci)
        c.fill = YF; c.border = T
        c.font = Font(bold=True, size=10); c.alignment = CC
    ws[f"A{tr}"].value = "TOTAL"
    ws[f"H{tr}"].value = f"=SUM(H7:H{tr-1})"
    ws[f"J{tr}"].value = f"=SUM(J7:J{tr-1})"
    for ci, w in enumerate([6, 44, 20, 6, 12, 12, 12, 14, 14, 14, 38], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Heading & Lane Assignments ─────────────────────────────────
    ws2 = wb.create_sheet("Heading & Lane Assignments")
    for ci, h in enumerate([
        "Marker #", "Name", "Code", "Lat", "Lon",
        "Heading°", "Src", "Curvature",
        "Lanes Used", "Road Width", "Sep Width",
        "Heading Override?", "Lane Override?", "Notes"
    ], 1):
        hc(ws2, 1, ci, h)
    ws2.row_dimensions[1].height = 30

    for ri, pg in enumerate(polygons, 2):
        mk  = pg.marker
        ov  = per_marker_overrides.get(mk.index)
        cfg = spec.resolve(ov)
        fill = OF if (mk.index in per_marker_headings or ov is not None) else (AF if ri%2==0 else WF)
        hov  = str(per_marker_headings.get(mk.index, "—"))
        lov  = str(ov.num_lanes) if ov and ov.num_lanes else "—"
        lane_lbl = LANE_PRESETS.get(pg.num_lanes_used, {}).get("label", f"{pg.num_lanes_used}-Lane")
        for ci, v in enumerate([
            mk.index, mk.name, mk.placement_code or "—",
            round(mk.lat, 7), round(mk.lon, 7),
            round(pg.heading_deg, 2), pg.heading_source,
            pg.road_curvature.replace('_', ' ').title(),
            lane_lbl,
            round(cfg["road_width_m"], 2),
            round(cfg["separator_width_m"], 2),
            f"YES {hov}°" if mk.index in per_marker_headings else "no",
            f"YES {lov}-lane" if ov and ov.num_lanes else "no",
            "OVERRIDDEN" if (mk.index in per_marker_headings or ov) else "default",
        ], 1):
            dc(ws2, ri, ci, v, fill)
    for i in range(1, 15):
        ws2.column_dimensions[get_column_letter(i)].width = 16

    # ── Sheet 3: Marker Details ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Marker Details")
    for ci, h in enumerate([
        "#", "Name", "Code", "Lat", "Lon", "Heading°", "Src",
        "Curvature", "Lanes", "Strips", "Per Lane",
        "Strip W(mm)", "Road W(m)", "Lane W(m)", "Area/Strip"
    ], 1):
        hc(ws3, 1, ci, h)
    for ri, pg in enumerate(polygons, 2):
        mk  = pg.marker
        ov  = per_marker_overrides.get(mk.index)
        cfg = spec.resolve(ov)
        fill = AF if ri % 2 == 0 else WF
        lw2  = cfg["lane_width_m"]
        for ci, v in enumerate([
            mk.index, mk.name, mk.placement_code or "—",
            round(mk.lat, 7), round(mk.lon, 7),
            round(pg.heading_deg, 2), pg.heading_source,
            pg.road_curvature.replace('_', ' ').title(),
            pg.num_lanes_used, spec.num_strips,
            spec.num_strips // pg.num_lanes_used,
            spec.strip_width_mm, cfg["road_width_m"],
            round(lw2, 3), round(spec.strip_width_mm / 1000 * lw2, 5),
        ], 1):
            dc(ws3, ri, ci, v, fill)
    for i in range(1, 16):
        ws3.column_dimensions[get_column_letter(i)].width = 16

    # ── Sheet 4: Strip Coordinates ───────────────────────────────────────────
    ws4 = wb.create_sheet("Strip Coordinates")
    for ci, h in enumerate(["Marker#", "Name", "Lane", "Strip", "Corner", "Lon", "Lat"], 1):
        hc(ws4, 1, ci, h)
    r = 2
    for pg in polygons:
        for si, (strip, ln) in enumerate(zip(pg.strip_polygons, pg.lane_assignments), 1):
            fill = AF if si % 2 == 0 else WF
            for ci2, (lo, la) in enumerate(strip, 1):
                for ci3, v in enumerate([
                    pg.marker.index, pg.marker.name,
                    ln, si, ci2, round(lo, 8), round(la, 8)
                ], 1):
                    dc(ws4, r, ci3, v, fill)
                r += 1
    for i in range(1, 8):
        ws4.column_dimensions[get_column_letter(i)].width = 20

    # ── Sheet 5: Project Spec ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Project Spec")
    ws5.merge_cells("A1:B1")
    ws5['A1'].value = "Project Specification — Speed Breaker Installation"
    ws5['A1'].font = Font(bold=True, size=11, color="FFFFFF")
    ws5['A1'].fill = HF; ws5['A1'].alignment = CC; ws5['A1'].border = T

    ov_count = sum(1 for pg in polygons if pg.heading_source == "per-marker")
    lc_count = sum(1 for pg in polygons if pg.override and pg.override.num_lanes)
    rows = [
        ("Material",               "Cold Applied Plastic (CAP) PTBM — Parabolic Transverse Bar"),
        ("Strip Width",            f"{spec.strip_width_mm} mm"),
        ("Number of Strips",       str(spec.num_strips)),
        ("Gap Between Strips",     f"{spec.gap_between_strips_m} m"),
        ("Default Lanes",          f"{spec.num_lanes}"),
        ("Default Road Width",     f"{spec.road_width_m} m"),
        ("Default Separator",      f"{spec.separator_width_m} m"),
        ("Global Heading",         f"{spec.heading_override}° ({'manual' if spec.heading_override>=0 else 'auto'})"),
        ("Total Markers",          str(len(polygons))),
        ("Per-Marker Hdg Overrides", str(ov_count)),
        ("Per-Marker Lane Overrides", str(lc_count)),
        ("Total Strips Generated", str(sum(len(pg.strip_polygons) for pg in polygons))),
    ]
    for ri2, (lbl, val) in enumerate(rows, 2):
        fill = AF if ri2 % 2 == 0 else WF
        c1 = ws5.cell(row=ri2, column=1, value=lbl)
        c2 = ws5.cell(row=ri2, column=2, value=val)
        for c in [c1, c2]:
            c.fill = fill; c.border = T; c.alignment = LC
        c1.font = Font(bold=True, size=10)
        c2.font = Font(size=10)
    ws5.column_dimensions['A'].width = 32
    ws5.column_dimensions['B'].width = 52

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

def run_pipeline(
    kml_in:               str,
    kml_out:              str,
    xlsx_out:             str,
    spec:                 PolygonSpec,
    per_marker_headings:  Dict[int, float]       = None,
    per_marker_overrides: Dict[int, MarkerOverride] = None,
    use_osm:              bool                   = True,
    progress_callback                            = None,
) -> Tuple[List[MarkerInfo], List[GeneratedPolygon]]:

    from typing import Tuple
    per_marker_headings  = per_marker_headings  or {}
    per_marker_overrides = per_marker_overrides or {}

    markers = parse_kml_markers(kml_in)
    if not markers:
        raise ValueError("No point markers found in KML file.")

    pca_h    = pca_heading(markers)
    polygons = []

    for i, mk in enumerate(markers):
        if progress_callback:
            progress_callback(i, len(markers), mk.name)
        polygons.append(generate_polygon_for_marker(
            markers, i, spec, pca_h, use_osm,
            per_marker_headings, per_marker_overrides))

    export_kml(markers, polygons, kml_out, spec,
               per_marker_headings, per_marker_overrides)
    export_excel(polygons, xlsx_out, spec,
                 per_marker_headings, per_marker_overrides)
    return markers, polygons


from typing import Tuple