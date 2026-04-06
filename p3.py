# # """
# # p3.py v1 — Speed Breaker CAP PTBM Polygon Engine
# # IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI

# # NEW FEATURES:
# #   ● Google Maps Roads API — snaps markers to road, gets accurate per-point bearing
# #   ● Curve detection — polygon tilts to match local road angle on curved roads
# #   ● SLOW text overlay — added to KML as GroundOverlay near each speed breaker
# #   ● Excel with embedded satellite image per marker (Google Static Maps API)
# #   ● Strip VISIBLE WIDTH — 0.5m along-road thickness (clearly visible in satellite view)
# #   ● Accurate geodesic distance using Haversine (same as Google Earth)
# #   ● No Google Earth Pro needed — all preview in Streamlit
# # """
# # from __future__ import annotations
# # import math, os, io, xml.etree.ElementTree as ET, tempfile
# # from dataclasses import dataclass, field
# # from typing import Dict, List, Optional, Tuple
# # import requests
# # import openpyxl
# # from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# # from openpyxl.utils import get_column_letter
# # from openpyxl.drawing.image import Image as XLImage

# # R_EARTH = 6_371_000.0

# # LANE_PRESETS: Dict[str, dict] = {
# #     "1-Lane (3.5m)":  dict(num_lanes=1, road_width_m=3.5,  separator_width_m=0.0),
# #     "2-Lane (7.0m)":  dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
# #     "4-Lane (14.0m)": dict(num_lanes=4, road_width_m=14.0, separator_width_m=2.0),
# #     "6-Lane (21.0m)": dict(num_lanes=6, road_width_m=21.0, separator_width_m=3.0),
# #     "Custom":         dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
# # }

# # LANE_COLOURS_KML = ["ff00d7ff","ff0088ff","ff00ff88","ffff44aa","ff44ffcc","ffcc44ff"]


# # # ─────────────────────────────────────────────────────────────────
# # # Geometry helpers
# # # ─────────────────────────────────────────────────────────────────

# # def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
# #     p1, p2 = math.radians(lat1), math.radians(lat2)
# #     dp = math.radians(lat2 - lat1)
# #     dl = math.radians(lon2 - lon1)
# #     a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
# #     return 2 * R_EARTH * math.asin(math.sqrt(max(0.0, min(1.0, a))))


# # def forward_bearing(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
# #     p1, p2 = math.radians(lat1), math.radians(lat2)
# #     dl = math.radians(lon2 - lon1)
# #     x = math.sin(dl) * math.cos(p2)
# #     y = math.cos(p1)*math.sin(p2) - math.sin(p1)*math.cos(p2)*math.cos(dl)
# #     return (math.degrees(math.atan2(x, y)) + 360) % 360


# # def norm180(h: float) -> float:
# #     h = h % 360
# #     return h - 180 if h >= 180 else h


# # def offset_ll(lat: float, lon: float, bearing_deg: float, dist_m: float) -> Tuple[float, float]:
# #     d = dist_m / R_EARTH
# #     b = math.radians(bearing_deg)
# #     p1 = math.radians(lat)
# #     l1 = math.radians(lon)
# #     p2 = math.asin(math.sin(p1)*math.cos(d) + math.cos(p1)*math.sin(d)*math.cos(b))
# #     l2 = l1 + math.atan2(math.sin(b)*math.sin(d)*math.cos(p1),
# #                           math.cos(d) - math.sin(p1)*math.sin(p2))
# #     return math.degrees(p2), math.degrees(l2)


# # def build_strip_rect(
# #     centre_lat: float, centre_lon: float,
# #     road_heading: float,
# #     along_offset_m: float,
# #     road_width_m: float,
# #     strip_thick_m: float,           # thickness ALONG road direction (visible width)
# # ) -> List[Tuple[float, float]]:
# #     """
# #     Build ONE strip rectangle.
# #     - Spans road_width_m PERPENDICULAR to road (full road width)
# #     - strip_thick_m along road direction (0.5m default = clearly visible in satellite)
# #     Returns 5-point closed polygon.
# #     """
# #     # Move to the strip's along-road position
# #     sc_lat, sc_lon = offset_ll(centre_lat, centre_lon, road_heading, along_offset_m)
# #     perp = (road_heading + 90) % 360
# #     half_w = road_width_m / 2
# #     half_t = strip_thick_m / 2

# #     # Left and right edges (perpendicular to road)
# #     left_lat,  left_lon  = offset_ll(sc_lat, sc_lon, perp,  +half_w)
# #     right_lat, right_lon = offset_ll(sc_lat, sc_lon, perp,  -half_w)

# #     # 4 corners (forward/backward along road from each edge)
# #     lf = offset_ll(left_lat,  left_lon,  road_heading, +half_t)
# #     lb = offset_ll(left_lat,  left_lon,  road_heading, -half_t)
# #     rb = offset_ll(right_lat, right_lon, road_heading, -half_t)
# #     rf = offset_ll(right_lat, right_lon, road_heading, +half_t)
# #     return [lf, lb, rb, rf, lf]


# # # ─────────────────────────────────────────────────────────────────
# # # Google Maps API helpers
# # # ─────────────────────────────────────────────────────────────────

# # def snap_to_roads(
# #     markers: List["KMLMarker"],
# #     api_key: str,
# # ) -> Dict[int, Tuple[float, float]]:
# #     """
# #     Use Google Maps Roads API to snap each marker to nearest road.
# #     Returns {marker_index: (snapped_lat, snapped_lon)}.
# #     Falls back to original coords if API fails.
# #     """
# #     if not api_key or not markers:
# #         return {}

# #     # Build path string (max 100 points per request)
# #     snapped: Dict[int, Tuple[float, float]] = {}

# #     chunk_size = 100
# #     for chunk_start in range(0, len(markers), chunk_size):
# #         chunk = markers[chunk_start:chunk_start + chunk_size]
# #         path = "|".join(f"{m.lat},{m.lon}" for m in chunk)
# #         url = "https://roads.googleapis.com/v1/snapToRoads"
# #         params = {
# #             "path": path,
# #             "interpolate": "false",
# #             "key": api_key,
# #         }
# #         try:
# #             resp = requests.get(url, params=params, timeout=10)
# #             data = resp.json()
# #             if "snappedPoints" in data:
# #                 for pt in data["snappedPoints"]:
# #                     orig_idx = pt.get("originalIndex", -1)
# #                     if orig_idx >= 0:
# #                         loc = pt["location"]
# #                         abs_idx = chunk[orig_idx].index
# #                         snapped[abs_idx] = (loc["latitude"], loc["longitude"])
# #         except Exception:
# #             pass  # Fall back to original coords

# #     return snapped


# # def get_road_heading_from_maps(
# #     lat: float, lon: float, api_key: str, radius_m: float = 30
# # ) -> Optional[float]:
# #     """
# #     Query Nearest Roads to get accurate bearing at a point.
# #     Uses two nearby snapped points to compute heading.
# #     """
# #     if not api_key:
# #         return None
# #     # Try to get 2 nearby points along road to determine direction
# #     # Use a small offset and snap both
# #     offsets = [
# #         offset_ll(lat, lon, 0, radius_m),    # North offset
# #         offset_ll(lat, lon, 90, radius_m),   # East offset
# #     ]
# #     path = f"{lat},{lon}|{offsets[0][0]},{offsets[0][1]}|{offsets[1][0]},{offsets[1][1]}"
# #     url = "https://roads.googleapis.com/v1/snapToRoads"
# #     params = {"path": path, "interpolate": "true", "key": api_key}
# #     try:
# #         resp = requests.get(url, params=params, timeout=8)
# #         data = resp.json()
# #         pts = data.get("snappedPoints", [])
# #         if len(pts) >= 2:
# #             p1 = pts[0]["location"]
# #             p2 = pts[1]["location"]
# #             return norm180(forward_bearing(
# #                 p1["latitude"], p1["longitude"],
# #                 p2["latitude"], p2["longitude"]
# #             ))
# #     except Exception:
# #         pass
# #     return None


# # def get_satellite_image(
# #     lat: float, lon: float,
# #     api_key: str,
# #     zoom: int = 20,
# #     size: str = "400x300",
# #     marker_color: str = "red",
# # ) -> Optional[bytes]:
# #     """
# #     Fetch a Google Static Maps satellite image centered at (lat, lon).
# #     Returns PNG bytes or None on failure.
# #     """
# #     if not api_key:
# #         return None
# #     url = "https://maps.googleapis.com/maps/api/staticmap"
# #     params = {
# #         "center":  f"{lat},{lon}",
# #         "zoom":    str(zoom),
# #         "size":    size,
# #         "maptype": "satellite",
# #         "markers": f"color:{marker_color}|{lat},{lon}",
# #         "key":     api_key,
# #     }
# #     try:
# #         resp = requests.get(url, params=params, timeout=15)
# #         if resp.status_code == 200 and resp.headers.get("content-type","").startswith("image"):
# #             return resp.content
# #     except Exception:
# #         pass
# #     return None


# # # ─────────────────────────────────────────────────────────────────
# # # KML parsing
# # # ─────────────────────────────────────────────────────────────────

# # @dataclass
# # class KMLMarker:
# #     name: str
# #     lat:  float
# #     lon:  float
# #     index: int = 0


# # def parse_kml(path: str) -> List[KMLMarker]:
# #     """Parse KML — returns only Point placemarks (center markers)."""
# #     root = ET.parse(path).getroot()
# #     markers: List[KMLMarker] = []

# #     def iter_pm(node):
# #         for c in node:
# #             if c.tag.split("}")[-1] == "Placemark": yield c
# #             else: yield from iter_pm(c)

# #     idx = 0
# #     for pm in iter_pm(root):
# #         name = f"Marker_{idx+1}"
# #         for c in pm:
# #             if c.tag.split("}")[-1] == "name":
# #                 name = (c.text or "").strip() or name
# #                 break
# #         for el in pm.iter():
# #             if el.tag.split("}")[-1] == "Point":
# #                 for ce in el.iter():
# #                     if ce.tag.split("}")[-1] == "coordinates" and ce.text:
# #                         p = ce.text.strip().split(",")
# #                         if len(p) >= 2:
# #                             try:
# #                                 markers.append(KMLMarker(name, float(p[1]), float(p[0]), idx))
# #                                 idx += 1
# #                             except ValueError:
# #                                 pass
# #                 break
# #     return markers


# # # ─────────────────────────────────────────────────────────────────
# # # Heading detection (neighbours + optional Roads API)
# # # ─────────────────────────────────────────────────────────────────

# # def detect_heading(
# #     markers: List[KMLMarker],
# #     idx: int,
# #     api_key: str = "",
# #     snapped: Optional[Dict[int, Tuple[float, float]]] = None,
# #     window: int = 3,
# # ) -> Tuple[float, str]:
# #     """
# #     Detect road heading at marker idx.
# #     Priority:
# #       1. Google Roads API (most accurate for curves)
# #       2. Snapped neighbour bearings
# #       3. Raw neighbour bearings
# #     """
# #     mk = markers[idx]
# #     lat, lon = mk.lat, mk.lon

# #     # Use snapped position if available
# #     if snapped and idx in snapped:
# #         lat, lon = snapped[idx]

# #     # Try Roads API first
# #     if api_key:
# #         h = get_road_heading_from_maps(lat, lon, api_key)
# #         if h is not None:
# #             return h, "Google Roads API"

# #     # Neighbour-based detection using snapped positions
# #     bearings: List[Tuple[float, float]] = []
# #     for offset in range(-window, window + 1):
# #         if offset == 0: continue
# #         j = idx + offset
# #         if j < 0 or j >= len(markers): continue
# #         nb = markers[j]
# #         nb_lat = lat  # same marker position for self
# #         nb_lat, nb_lon = nb.lat, nb.lon
# #         # Use snapped position for neighbour if available
# #         if snapped and j in snapped:
# #             nb_lat, nb_lon = snapped[j]
# #         d = haversine(lat, lon, nb_lat, nb_lon)
# #         if d < 0.5: continue
# #         b = norm180(forward_bearing(lat, lon, nb_lat, nb_lon))
# #         w = 1.0 / (abs(offset) * max(d, 1.0))
# #         bearings.append((b, w))

# #     if not bearings:
# #         return 0.0, "default (0°)"

# #     sx = sum(w * math.cos(math.radians(2*b)) for b, w in bearings)
# #     sy = sum(w * math.sin(math.radians(2*b)) for b, w in bearings)
# #     return norm180(math.degrees(math.atan2(sy, sx)) / 2), "neighbour-avg"


# # # ─────────────────────────────────────────────────────────────────
# # # Spec
# # # ─────────────────────────────────────────────────────────────────

# # @dataclass
# # class PolySpec:
# #     road_width_m:      float = 7.0
# #     num_lanes:         int   = 2
# #     separator_width_m: float = 0.5
# #     num_strips:        int   = 3
# #     # strip_thick_m: thickness ALONG road (visible width in satellite)
# #     # 0.5m default = clearly visible as rectangle (not a thin line)
# #     strip_thick_m:     float = 0.5
# #     strip_gap_m:       float = 0.6    # gap between strips along road
# #     heading_override:  Optional[float] = None
# #     add_slow_label:    bool  = True   # add SLOW text to KML
# #     api_key:           str   = ""
# #     marker_overrides:  Dict[int, dict] = field(default_factory=dict)


# # # ─────────────────────────────────────────────────────────────────
# # # Generated polygon
# # # ─────────────────────────────────────────────────────────────────

# # @dataclass
# # class GenPoly:
# #     marker_idx:    int
# #     marker_name:   str
# #     strip_idx:     int
# #     coords:        List[Tuple[float, float]]
# #     road_heading:  float
# #     heading_src:   str
# #     road_width_m:  float
# #     lane_width_m:  float
# #     strip_thick_m: float
# #     strip_gap_m:   float
# #     num_lanes:     int
# #     along_offset_m: float
# #     snapped_lat:   float
# #     snapped_lon:   float


# # # ─────────────────────────────────────────────────────────────────
# # # Main generator
# # # ─────────────────────────────────────────────────────────────────

# # def generate_polygons(
# #     markers: List[KMLMarker],
# #     spec: PolySpec,
# #     snapped: Optional[Dict[int, Tuple[float, float]]] = None,
# # ) -> Tuple[List[GenPoly], Dict[int, Tuple[float, float, str]]]:
# #     """
# #     Generate strip polygons for all markers.
# #     Returns (all_polys, headings_dict)
# #     headings_dict: {marker_idx: (heading_deg, snap_lat, snap_lon, src_label)}
# #     """
# #     all_polys: List[GenPoly] = []
# #     headings: Dict[int, Tuple[float, float, float, str]] = {}

# #     for mk in markers:
# #         i = mk.index
# #         ov = spec.marker_overrides.get(i, {})

# #         rw  = float(ov.get("road_width_m",  spec.road_width_m))
# #         nl  = int(ov.get("num_lanes",        spec.num_lanes))
# #         sep = float(ov.get("separator_width_m", spec.separator_width_m))
# #         ns  = int(ov.get("num_strips",       spec.num_strips))
# #         st  = float(ov.get("strip_thick_m",  spec.strip_thick_m))
# #         gap = float(ov.get("strip_gap_m",    spec.strip_gap_m))
# #         lw  = (rw - sep) / max(nl, 1)

# #         # Snapped centre
# #         snap_lat = mk.lat
# #         snap_lon = mk.lon
# #         if snapped and i in snapped:
# #             snap_lat, snap_lon = snapped[i]

# #         # Heading
# #         manual_h = ov.get("heading_deg", spec.heading_override)
# #         if manual_h is not None:
# #             heading = norm180(float(manual_h))
# #             hsrc    = "manual"
# #         else:
# #             heading, hsrc = detect_heading(markers, i, spec.api_key, snapped)

# #         headings[i] = (heading, snap_lat, snap_lon, hsrc)

# #         # Strip positions along road (centred on marker)
# #         total_span = ns * st + (ns - 1) * gap
# #         first_pos  = -total_span / 2 + st / 2

# #         for si in range(ns):
# #             along = first_pos + si * (st + gap)
# #             coords = build_strip_rect(snap_lat, snap_lon, heading, along, rw, st)
# #             all_polys.append(GenPoly(
# #                 marker_idx    = i,
# #                 marker_name   = mk.name,
# #                 strip_idx     = si,
# #                 coords        = coords,
# #                 road_heading  = heading,
# #                 heading_src   = hsrc,
# #                 road_width_m  = rw,
# #                 lane_width_m  = lw,
# #                 strip_thick_m = st,
# #                 strip_gap_m   = gap,
# #                 num_lanes     = nl,
# #                 along_offset_m = along,
# #                 snapped_lat   = snap_lat,
# #                 snapped_lon   = snap_lon,
# #             ))

# #     return all_polys, headings


# # # ─────────────────────────────────────────────────────────────────
# # # Pipeline
# # # ─────────────────────────────────────────────────────────────────

# # def run_pipeline(
# #     kml_path: str,
# #     spec: PolySpec,
# #     per_headings: Optional[Dict[int, float]] = None,
# #     progress_cb=None,  # optional callback(step: str, pct: int)
# # ) -> Tuple[List[KMLMarker], List[GenPoly], Dict]:
# #     """
# #     Full pipeline:
# #       1. Parse KML
# #       2. Snap to roads (if API key)
# #       3. Detect headings
# #       4. Generate polygons
# #     """
# #     def _cb(msg, pct):
# #         if progress_cb: progress_cb(msg, pct)

# #     _cb("Parsing KML…", 10)
# #     markers = parse_kml(kml_path)
# #     if not markers:
# #         raise ValueError("No Point markers found in KML.")

# #     # Apply per-marker heading overrides
# #     if per_headings:
# #         for idx, hdg in per_headings.items():
# #             spec.marker_overrides.setdefault(idx, {})["heading_deg"] = hdg

# #     _cb("Snapping to roads via Google API…", 30)
# #     snapped: Dict[int, Tuple[float, float]] = {}
# #     if spec.api_key:
# #         snapped = snap_to_roads(markers, spec.api_key)

# #     _cb("Detecting road headings…", 50)
# #     all_polys, headings = generate_polygons(markers, spec, snapped)

# #     _cb("Done", 100)
# #     return markers, all_polys, headings


# # # ─────────────────────────────────────────────────────────────────
# # # KML export (with SLOW GroundOverlay label)
# # # ─────────────────────────────────────────────────────────────────

# # # SLOW label as a simple text polygon using thin lines spelling "SLOW"
# # # We use a Placemark with a Label style instead of complex text rendering

# # def _slow_label_kml(
# #     snap_lat: float, snap_lon: float,
# #     heading: float,
# #     road_width_m: float,
# #     strip_thick_m: float,
# #     ahead_offset_m: float = 2.0,  # place SLOW label this far ahead of strips
# # ) -> str:
# #     """
# #     Generate KML for a 'SLOW' text marker positioned ahead of the speed breaker.
# #     Uses a visible Placemark pin with label style.
# #     """
# #     # Position label ahead of speed breaker
# #     label_lat, label_lon = offset_ll(snap_lat, snap_lon, heading, ahead_offset_m + strip_thick_m)

# #     return (
# #         f'<Placemark>'
# #         f'<n>SLOW</n>'
# #         f'<description>Speed breaker ahead warning</description>'
# #         f'<styleUrl>#slow_label</styleUrl>'
# #         f'<Point><coordinates>{label_lon:.8f},{label_lat:.8f},0</coordinates></Point>'
# #         f'</Placemark>'
# #     )


# # def export_kml(
# #     markers:   List[KMLMarker],
# #     all_polys: List[GenPoly],
# #     headings:  Dict,
# #     spec:      PolySpec,
# #     out_path:  str,
# # ) -> None:
# #     def cs(coords):
# #         return " ".join(f"{lo:.8f},{la:.8f},0" for la, lo in coords)

# #     by: Dict[int, List[GenPoly]] = {}
# #     for p in all_polys: by.setdefault(p.marker_idx, []).append(p)

# #     lines = [
# #         '<?xml version="1.0" encoding="UTF-8"?>',
# #         '<kml xmlns="http://www.opengis.net/kml/2.2">',
# #         '<Document>',
# #         '<n>CAP PTBM Speed Breaker Polygons — GIS BOQ v1 (p3)</n>',
# #         # ── Styles ──
# #         # Yellow fill polygon (visible rectangle)
# #         '<Style id="strip_yellow">',
# #         '  <LineStyle><color>ff1400ff</color><width>1</width></LineStyle>',
# #         '  <PolyStyle><color>cc00d7ff</color><fill>1</fill><outline>1</outline></PolyStyle>',
# #         '</Style>',
# #         # Orange fill for alternate strips
# #         '<Style id="strip_orange">',
# #         '  <LineStyle><color>ff1400ff</color><width>1</width></LineStyle>',
# #         '  <PolyStyle><color>cc0088ff</color><fill>1</fill><outline>1</outline></PolyStyle>',
# #         '</Style>',
# #         # Red pushpin for center marker
# #         '<Style id="pin">',
# #         '  <IconStyle><scale>1.1</scale>',
# #         '    <Icon><href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href></Icon>',
# #         '  </IconStyle>',
# #         '  <LabelStyle><scale>0.9</scale></LabelStyle>',
# #         '</Style>',
# #         # SLOW text label style
# #         '<Style id="slow_label">',
# #         '  <IconStyle><scale>0</scale>',
# #         '    <Icon><href>http://maps.google.com/mapfiles/kml/shapes/arrow.png</href></Icon>',
# #         '  </IconStyle>',
# #         '  <LabelStyle><color>ff00ffff</color><scale>1.2</scale></LabelStyle>',
# #         '</Style>',
# #     ]

# #     for mk in markers:
# #         ps = by.get(mk.index, [])
# #         h_data = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
# #         heading   = h_data[0]
# #         snap_lat  = h_data[1]
# #         snap_lon  = h_data[2]
# #         hsrc      = h_data[3]
# #         rw        = ps[0].road_width_m if ps else spec.road_width_m
# #         nl        = ps[0].num_lanes    if ps else spec.num_lanes
# #         st        = ps[0].strip_thick_m if ps else spec.strip_thick_m
# #         sw_mm_val = st * 1000

# #         lines.append(f'<Folder><n>{mk.name}</n>')

# #         # Center marker pin
# #         lines.append(
# #             f'<Placemark><n>{mk.name}</n>'
# #             f'<description>CAP PTBM {sw_mm_val:.0f}MM X {len(ps)}'
# #             f'&#10;Road: {rw:.1f}m | {nl} Lane'
# #             f'&#10;Heading: {heading:.1f}° [{hsrc}]'
# #             f'&#10;Lat: {snap_lat:.6f} Lon: {snap_lon:.6f}</description>'
# #             f'<styleUrl>#pin</styleUrl>'
# #             f'<Point><coordinates>{snap_lon:.8f},{snap_lat:.8f},0</coordinates></Point>'
# #             f'</Placemark>'
# #         )

# #         # Strip polygons — bright yellow filled rectangles
# #         for p in ps:
# #             style_id = "strip_yellow" if p.strip_idx % 2 == 0 else "strip_orange"
# #             label = f"CAP PTBM {p.strip_thick_m*1000:.0f}MM X {len(ps)}"
# #             lines.append(
# #                 f'<Placemark><n>{label}</n>'
# #                 f'<description>{mk.name} Strip {p.strip_idx+1}/{len(ps)}'
# #                 f'&#10;Width (road): {p.road_width_m:.2f}m'
# #                 f'&#10;Thick (along road): {p.strip_thick_m*1000:.0f}mm'
# #                 f'&#10;Heading: {p.road_heading:.1f}°</description>'
# #                 f'<styleUrl>#{style_id}</styleUrl>'
# #                 f'<Polygon><outerBoundaryIs><LinearRing>'
# #                 f'<coordinates>{cs(p.coords)}</coordinates>'
# #                 f'</LinearRing></outerBoundaryIs></Polygon>'
# #                 f'</Placemark>'
# #             )

# #         # SLOW label (if enabled)
# #         if spec.add_slow_label and ps:
# #             lines.append(_slow_label_kml(snap_lat, snap_lon, heading, rw, st))

# #         lines.append('</Folder>')

# #     lines += ['</Document>', '</kml>']
# #     with open(out_path, "w", encoding="utf-8") as f:
# #         f.write("\n".join(lines))


# # # ─────────────────────────────────────────────────────────────────
# # # Excel export with embedded satellite images
# # # ─────────────────────────────────────────────────────────────────

# # def _tb() -> Border:
# #     s = Side(style="thin")
# #     return Border(left=s, right=s, top=s, bottom=s)

# # def _fill(h: str) -> PatternFill:
# #     return PatternFill("solid", start_color=h, end_color=h)

# # def _hdr(ws, r, c, v, bg="1F3864", fg="FFFFFF"):
# #     x = ws.cell(r, c, v)
# #     x.font      = Font(bold=True, color=fg, size=9)
# #     x.fill      = _fill(bg)
# #     x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
# #     x.border    = _tb()
# #     return x

# # def _dat(ws, r, c, v, bg=None, bold=False):
# #     x = ws.cell(r, c, v)
# #     x.font      = Font(bold=bold, size=9)
# #     x.alignment = Alignment(horizontal="center", vertical="center")
# #     x.border    = _tb()
# #     if bg: x.fill = _fill(bg)
# #     return x


# # def export_excel(
# #     markers:   List[KMLMarker],
# #     all_polys: List[GenPoly],
# #     headings:  Dict,
# #     spec:      PolySpec,
# #     out_path:  str,
# #     progress_cb=None,
# # ) -> None:
# #     def _cb(msg, pct):
# #         if progress_cb: progress_cb(msg, pct)

# #     wb = openpyxl.Workbook()

# #     # ── Sheet 1: BOQ Summary (matching friend's format) ──────────────
# #     ws1 = wb.active
# #     ws1.title = "BOQ Summary"

# #     # Headers matching screenshot: S.No, Placemark, Lat, Lon, Bearing, Road Width,
# #     # Mark Length, Area, Rate, Amount, Satellite View
# #     H1 = [
# #         "S.No", "Placemark Name", "Latitude", "Longitude",
# #         "Bearing (°)", "Road Width (m)", "Mark Length (m)",
# #         "No. of Strips", "Area (m²)", "Rate (Rs/m²)", "Amount (Rs)",
# #         "Satellite View",
# #     ]
# #     for ci, h in enumerate(H1, 1):
# #         _hdr(ws1, 1, ci, h)
# #     ws1.row_dimensions[1].height = 36

# #     by: Dict[int, List[GenPoly]] = {}
# #     for p in all_polys: by.setdefault(p.marker_idx, []).append(p)

# #     # Row height for image rows
# #     IMAGE_ROW_H = 180  # pixels → approx 135pt

# #     for ri, mk in enumerate(markers, 2):
# #         ps    = by.get(mk.index, [])
# #         p0    = ps[0] if ps else None
# #         hdata = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
# #         h_deg     = hdata[0]
# #         snap_lat  = hdata[1]
# #         snap_lon  = hdata[2]
# #         rw        = p0.road_width_m  if p0 else spec.road_width_m
# #         nl        = p0.num_lanes     if p0 else spec.num_lanes
# #         st        = p0.strip_thick_m if p0 else spec.strip_thick_m
# #         gap       = p0.strip_gap_m   if p0 else spec.strip_gap_m
# #         ns        = len(ps)
# #         # Mark length = total span of all strips along road
# #         mark_len  = ns * st + max(ns-1, 0) * gap
# #         area_one  = st * rw
# #         area_tot  = area_one * ns
# #         rate      = 2500  # Rs/m² default (can be changed)
# #         amount    = round(area_tot * rate, 2)

# #         bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
# #         row_data = [
# #             ri - 1,
# #             mk.name,
# #             round(snap_lat, 6),
# #             round(snap_lon, 6),
# #             round(h_deg, 1),
# #             round(rw, 2),
# #             round(mark_len, 3),
# #             ns,
# #             round(area_tot, 4),
# #             rate,
# #             amount,
# #             "",  # Satellite View — image inserted separately
# #         ]
# #         for ci, val in enumerate(row_data, 1):
# #             _dat(ws1, ri, ci, val, bg=bg)

# #         # Set tall row for image
# #         ws1.row_dimensions[ri].height = IMAGE_ROW_H

# #         # Fetch and embed satellite image
# #         if spec.api_key:
# #             _cb(f"Fetching satellite image for {mk.name}…", 70)
# #             img_bytes = get_satellite_image(snap_lat, snap_lon, spec.api_key,
# #                                             zoom=20, size="400x300")
# #             if img_bytes:
# #                 try:
# #                     img_stream = io.BytesIO(img_bytes)
# #                     xl_img = XLImage(img_stream)
# #                     xl_img.width  = 260
# #                     xl_img.height = 180
# #                     # Column L = 12, row ri
# #                     cell_ref = f"L{ri}"
# #                     ws1.add_image(xl_img, cell_ref)
# #                 except Exception:
# #                     ws1.cell(ri, 12, "Image unavailable")
# #         else:
# #             ws1.cell(ri, 12, "Add GOOGLE_API_KEY for satellite images")

# #     # Column widths
# #     for i, w in enumerate([5, 24, 12, 12, 10, 13, 14, 10, 12, 12, 14, 45], 1):
# #         ws1.column_dimensions[get_column_letter(i)].width = w
# #     # Image column extra wide
# #     ws1.column_dimensions["L"].width = 45

# #     # ── Sheet 2: Strip Coordinates ──────────────────────────────────
# #     ws2 = wb.create_sheet("Strip Coordinates")
# #     H2 = ["S.No", "Marker", "Strip #", "Along Offset (m)",
# #           "C1 Lat", "C1 Lon", "C2 Lat", "C2 Lon",
# #           "C3 Lat", "C3 Lon", "C4 Lat", "C4 Lon",
# #           "Road Width (m)", "Thick (mm)", "Gap (m)", "Heading °"]
# #     for ci, h in enumerate(H2, 1): _hdr(ws2, 1, ci, h, bg="145A32")
# #     ws2.row_dimensions[1].height = 36

# #     for gi, p in enumerate(all_polys, 1):
# #         ri = gi + 1
# #         bg = "E9F7EF" if ri % 2 == 0 else "FFFFFF"
# #         cs_ = p.coords[:4]
# #         row2 = [gi, p.marker_name, p.strip_idx+1, round(p.along_offset_m, 4),
# #                 *[round(v, 8) for c_ in cs_ for v in c_],
# #                 round(p.road_width_m, 3), round(p.strip_thick_m*1000, 1),
# #                 round(p.strip_gap_m, 3), round(p.road_heading, 1)]
# #         for ci, val in enumerate(row2, 1): _dat(ws2, ri, ci, val, bg=bg)
# #     for i, w in enumerate([5, 22, 8, 14, 12, 12, 12, 12, 12, 12, 12, 12, 12, 10, 9, 10], 1):
# #         ws2.column_dimensions[get_column_letter(i)].width = w

# #     # ── Sheet 3: Spec Summary ────────────────────────────────────────
# #     ws3 = wb.create_sheet("Spec Summary")
# #     _hdr(ws3, 1, 1, "Parameter", bg="784212")
# #     _hdr(ws3, 1, 2, "Value", bg="784212")
# #     ws3.column_dimensions["A"].width = 32
# #     ws3.column_dimensions["B"].width = 60
# #     spec_rows = [
# #         ("Tool",            "GIS BOQ Speed Breaker Tool v1 (p3)"),
# #         ("Institute",       "IIIT Nagpur — Dr. Neha Kasture"),
# #         ("Client",          "PWD / NHAI"),
# #         ("", ""),
# #         ("Strip Type",      "CAP PTBM (Capsule Prefab Thermoplastic Bituminous Marking)"),
# #         ("Strip Orientation","PERPENDICULAR to road — spans full road width"),
# #         ("Strip Thickness", f"{spec.strip_thick_m*1000:.0f} mm thick (along road) = {spec.strip_thick_m:.2f}m"),
# #         ("Road Width",      f"{spec.road_width_m:.1f} m (strip length across road)"),
# #         ("Number of Strips",str(spec.num_strips)),
# #         ("Gap Between",     f"{spec.strip_gap_m:.2f} m (along road)"),
# #         ("SLOW Label",      "Yes — added to KML 2m ahead of strips" if spec.add_slow_label else "No"),
# #         ("Google API",      "Connected" if spec.api_key else "Not set — using neighbour heading"),
# #         ("Heading Method",  "Google Roads API + snap-to-road" if spec.api_key else "Neighbour bearing average"),
# #         ("", ""),
# #         ("Total Markers",   str(len(markers))),
# #         ("Total Strips",    str(len(all_polys))),
# #         ("Rate (default)",  "Rs 2,500 / m²"),
# #         ("", ""),
# #         ("NOTE",            "All markers assumed to be at CENTER of road. Polygon spans ±(road_width/2) cross-road."),
# #     ]
# #     for ri2, (k, v) in enumerate(spec_rows, 2):
# #         ws3.cell(ri2, 1, k).font = Font(bold=bool(k), size=9)
# #         ws3.cell(ri2, 2, v).font = Font(size=9)

# #     wb.save(out_path)


# # # ─────────────────────────────────────────────────────────────────
# # # Self-test
# # # ─────────────────────────────────────────────────────────────────
# # if __name__ == "__main__":
# #     import tempfile, os
# #     TEST_KML = """<?xml version="1.0" encoding="UTF-8"?>
# # <kml xmlns="http://www.opengis.net/kml/2.2"><Document>
# #   <Placemark><n>SB_1</n><Point><coordinates>93.94313688,24.83678499,0</coordinates></Point></Placemark>
# #   <Placemark><n>SB_2</n><Point><coordinates>93.94334266,24.83687056,0</coordinates></Point></Placemark>
# #   <Placemark><n>SB_3</n><Point><coordinates>93.94342449,24.83672434,0</coordinates></Point></Placemark>
# # </Document></kml>"""
# #     with tempfile.NamedTemporaryFile(mode="w", suffix=".kml", delete=False) as f:
# #         f.write(TEST_KML); kp = f.name

# #     # Load API key from env if available
# #     api_key = ""
# #     env_path = r"D:\Road_Safety_Reasearch_III\.env"
# #     if os.path.exists(env_path):
# #         for line in open(env_path):
# #             if line.startswith("GOOGLE_API_KEY="):
# #                 api_key = line.split("=", 1)[1].strip()
# #                 break

# #     spec = PolySpec(
# #         road_width_m=7.0, num_lanes=2, separator_width_m=0.5,
# #         num_strips=3, strip_thick_m=0.5, strip_gap_m=0.6,
# #         add_slow_label=True, api_key=api_key,
# #     )
# #     markers, polys, headings = run_pipeline(kp, spec)
# #     print(f"\n✅ Markers: {len(markers)} | Strips: {len(polys)}")
# #     for mk in markers:
# #         ps = [p for p in polys if p.marker_idx == mk.index]
# #         hd = headings[mk.index]
# #         print(f"  📍 {mk.name}  snap=({hd[1]:.6f},{hd[2]:.6f})  "
# #               f"heading={hd[0]:.1f}° [{hd[3]}]  strips={len(ps)}")
# #         for p in ps:
# #             print(f"     Strip {p.strip_idx+1}: offset={p.along_offset_m:+.3f}m  "
# #                   f"road_w={p.road_width_m:.1f}m  thick={p.strip_thick_m*1000:.0f}mm")

# #     export_kml(markers, polys, headings, spec, "/tmp/p3_test.kml")
# #     export_excel(markers, polys, headings, spec, "/tmp/p3_test.xlsx")
# #     print("\n✅ KML + Excel exported")
# #     os.unlink(kp)


# """
# p3.py v2 — Speed Breaker CAP PTBM Polygon Engine
# IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI

# FEATURES:
#   ● Google Maps Roads API — snaps markers to road, gets accurate per-point bearing
#   ● Curve detection — polygon tilts to match local road angle on curved roads
#   ● SLOW text overlay — added to KML as GroundOverlay near each speed breaker
#   ● Excel with embedded satellite image per marker (Google Static Maps API)
#   ● Strip VISIBLE WIDTH — 0.5m along-road thickness (clearly visible in satellite view)
#   ● Accurate geodesic distance using Haversine (same as Google Earth)
#   ● No Google Earth Pro needed — all preview in Streamlit
# """
# from __future__ import annotations
# import math, os, io, xml.etree.ElementTree as ET, tempfile
# from dataclasses import dataclass, field
# from typing import Dict, List, Optional, Tuple
# import requests
# import openpyxl
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.drawing.image import Image as XLImage

# R_EARTH = 6_371_000.0

# LANE_PRESETS: Dict[str, dict] = {
#     "1-Lane (3.5m)":  dict(num_lanes=1, road_width_m=3.5,  separator_width_m=0.0),
#     "2-Lane (7.0m)":  dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
#     "4-Lane (14.0m)": dict(num_lanes=4, road_width_m=14.0, separator_width_m=2.0),
#     "6-Lane (21.0m)": dict(num_lanes=6, road_width_m=21.0, separator_width_m=3.0),
#     "Custom":         dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
# }

# LANE_COLOURS_KML = ["ff00d7ff","ff0088ff","ff00ff88","ffff44aa","ff44ffcc","ffcc44ff"]


# # ─────────────────────────────────────────────────────────────────
# # Geometry helpers
# # ─────────────────────────────────────────────────────────────────

# def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
#     p1, p2 = math.radians(lat1), math.radians(lat2)
#     dp = math.radians(lat2 - lat1)
#     dl = math.radians(lon2 - lon1)
#     a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
#     return 2 * R_EARTH * math.asin(math.sqrt(max(0.0, min(1.0, a))))


# def forward_bearing(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
#     p1, p2 = math.radians(lat1), math.radians(lat2)
#     dl = math.radians(lon2 - lon1)
#     x = math.sin(dl) * math.cos(p2)
#     y = math.cos(p1)*math.sin(p2) - math.sin(p1)*math.cos(p2)*math.cos(dl)
#     return (math.degrees(math.atan2(x, y)) + 360) % 360


# def norm180(h: float) -> float:
#     h = h % 360
#     return h - 180 if h >= 180 else h


# def offset_ll(lat: float, lon: float, bearing_deg: float, dist_m: float) -> Tuple[float, float]:
#     d = dist_m / R_EARTH
#     b = math.radians(bearing_deg)
#     p1 = math.radians(lat)
#     l1 = math.radians(lon)
#     p2 = math.asin(math.sin(p1)*math.cos(d) + math.cos(p1)*math.sin(d)*math.cos(b))
#     l2 = l1 + math.atan2(math.sin(b)*math.sin(d)*math.cos(p1),
#                           math.cos(d) - math.sin(p1)*math.sin(p2))
#     return math.degrees(p2), math.degrees(l2)


# def build_strip_rect(
#     centre_lat: float, centre_lon: float,
#     road_heading: float,
#     along_offset_m: float,
#     road_width_m: float,
#     strip_thick_m: float,
# ) -> List[Tuple[float, float]]:
#     """
#     Build ONE strip rectangle.
#     - Spans road_width_m PERPENDICULAR to road (full road width)
#     - strip_thick_m along road direction (visible width)
#     Returns 5-point closed polygon.
#     """
#     sc_lat, sc_lon = offset_ll(centre_lat, centre_lon, road_heading, along_offset_m)
#     perp = (road_heading + 90) % 360
#     half_w = road_width_m / 2
#     half_t = strip_thick_m / 2

#     left_lat,  left_lon  = offset_ll(sc_lat, sc_lon, perp,  +half_w)
#     right_lat, right_lon = offset_ll(sc_lat, sc_lon, perp,  -half_w)

#     lf = offset_ll(left_lat,  left_lon,  road_heading, +half_t)
#     lb = offset_ll(left_lat,  left_lon,  road_heading, -half_t)
#     rb = offset_ll(right_lat, right_lon, road_heading, -half_t)
#     rf = offset_ll(right_lat, right_lon, road_heading, +half_t)
#     return [lf, lb, rb, rf, lf]


# # ─────────────────────────────────────────────────────────────────
# # Google Maps API helpers
# # ─────────────────────────────────────────────────────────────────

# def snap_to_roads(
#     markers: List["KMLMarker"],
#     api_key: str,
# ) -> Dict[int, Tuple[float, float]]:
#     """Snap each marker to nearest road using Google Roads API."""
#     if not api_key or not markers:
#         return {}

#     snapped: Dict[int, Tuple[float, float]] = {}
#     chunk_size = 100
#     for chunk_start in range(0, len(markers), chunk_size):
#         chunk = markers[chunk_start:chunk_start + chunk_size]
#         path = "|".join(f"{m.lat},{m.lon}" for m in chunk)
#         url = "https://roads.googleapis.com/v1/snapToRoads"
#         params = {"path": path, "interpolate": "false", "key": api_key}
#         try:
#             resp = requests.get(url, params=params, timeout=10)
#             data = resp.json()
#             if "snappedPoints" in data:
#                 for pt in data["snappedPoints"]:
#                     orig_idx = pt.get("originalIndex", -1)
#                     if orig_idx >= 0:
#                         loc = pt["location"]
#                         abs_idx = chunk[orig_idx].index
#                         snapped[abs_idx] = (loc["latitude"], loc["longitude"])
#         except Exception:
#             pass

#     return snapped


# def get_road_heading_from_maps(
#     lat: float, lon: float, api_key: str, radius_m: float = 30
# ) -> Optional[float]:
#     """Query Roads API to get accurate bearing at a point."""
#     if not api_key:
#         return None
#     offsets = [
#         offset_ll(lat, lon, 0, radius_m),
#         offset_ll(lat, lon, 90, radius_m),
#     ]
#     path = f"{lat},{lon}|{offsets[0][0]},{offsets[0][1]}|{offsets[1][0]},{offsets[1][1]}"
#     url = "https://roads.googleapis.com/v1/snapToRoads"
#     params = {"path": path, "interpolate": "true", "key": api_key}
#     try:
#         resp = requests.get(url, params=params, timeout=8)
#         data = resp.json()
#         pts = data.get("snappedPoints", [])
#         if len(pts) >= 2:
#             p1 = pts[0]["location"]
#             p2 = pts[1]["location"]
#             return norm180(forward_bearing(
#                 p1["latitude"], p1["longitude"],
#                 p2["latitude"], p2["longitude"]
#             ))
#     except Exception:
#         pass
#     return None


# def get_satellite_image(
#     lat: float, lon: float,
#     api_key: str,
#     zoom: int = 20,
#     size: str = "640x480",
#     marker_color: str = "red",
# ) -> Optional[bytes]:
#     """Fetch Google Static Maps satellite image at high zoom."""
#     if not api_key:
#         return None
#     url = "https://maps.googleapis.com/maps/api/staticmap"
#     params = {
#         "center":  f"{lat},{lon}",
#         "zoom":    str(zoom),
#         "size":    size,
#         "maptype": "satellite",
#         "markers": f"color:{marker_color}|size:small|{lat},{lon}",
#         "key":     api_key,
#         "scale":   "2",   # retina/2x for sharper image
#     }
#     try:
#         resp = requests.get(url, params=params, timeout=15)
#         if resp.status_code == 200 and resp.headers.get("content-type", "").startswith("image"):
#             return resp.content
#     except Exception:
#         pass
#     return None


# # ─────────────────────────────────────────────────────────────────
# # KML parsing
# # ─────────────────────────────────────────────────────────────────

# @dataclass
# class KMLMarker:
#     name: str
#     lat:  float
#     lon:  float
#     index: int = 0


# def parse_kml(path: str) -> List[KMLMarker]:
#     """Parse KML — returns only Point placemarks (center markers)."""
#     root = ET.parse(path).getroot()
#     markers: List[KMLMarker] = []

#     def iter_pm(node):
#         for c in node:
#             if c.tag.split("}")[-1] == "Placemark": yield c
#             else: yield from iter_pm(c)

#     idx = 0
#     for pm in iter_pm(root):
#         name = f"Marker_{idx+1}"
#         for c in pm:
#             if c.tag.split("}")[-1] == "name":
#                 name = (c.text or "").strip() or name
#                 break
#         for el in pm.iter():
#             if el.tag.split("}")[-1] == "Point":
#                 for ce in el.iter():
#                     if ce.tag.split("}")[-1] == "coordinates" and ce.text:
#                         p = ce.text.strip().split(",")
#                         if len(p) >= 2:
#                             try:
#                                 markers.append(KMLMarker(name, float(p[1]), float(p[0]), idx))
#                                 idx += 1
#                             except ValueError:
#                                 pass
#                 break
#     return markers


# # ─────────────────────────────────────────────────────────────────
# # Heading detection
# # ─────────────────────────────────────────────────────────────────

# def detect_heading(
#     markers: List[KMLMarker],
#     idx: int,
#     api_key: str = "",
#     snapped: Optional[Dict[int, Tuple[float, float]]] = None,
#     window: int = 3,
# ) -> Tuple[float, str]:
#     """Detect road heading at marker idx."""
#     mk = markers[idx]
#     lat, lon = mk.lat, mk.lon

#     if snapped and idx in snapped:
#         lat, lon = snapped[idx]

#     if api_key:
#         h = get_road_heading_from_maps(lat, lon, api_key)
#         if h is not None:
#             return h, "Google Roads API"

#     bearings: List[Tuple[float, float]] = []
#     for offset in range(-window, window + 1):
#         if offset == 0: continue
#         j = idx + offset
#         if j < 0 or j >= len(markers): continue
#         nb = markers[j]
#         nb_lat, nb_lon = nb.lat, nb.lon
#         if snapped and j in snapped:
#             nb_lat, nb_lon = snapped[j]
#         d = haversine(lat, lon, nb_lat, nb_lon)
#         if d < 0.5: continue
#         b = norm180(forward_bearing(lat, lon, nb_lat, nb_lon))
#         w = 1.0 / (abs(offset) * max(d, 1.0))
#         bearings.append((b, w))

#     if not bearings:
#         return 0.0, "default (0°)"

#     sx = sum(w * math.cos(math.radians(2*b)) for b, w in bearings)
#     sy = sum(w * math.sin(math.radians(2*b)) for b, w in bearings)
#     return norm180(math.degrees(math.atan2(sy, sx)) / 2), "neighbour-avg"


# # ─────────────────────────────────────────────────────────────────
# # Spec
# # ─────────────────────────────────────────────────────────────────

# @dataclass
# class PolySpec:
#     road_width_m:      float = 7.0
#     num_lanes:         int   = 2
#     separator_width_m: float = 0.5
#     num_strips:        int   = 3
#     strip_thick_m:     float = 0.5
#     strip_gap_m:       float = 0.6
#     heading_override:  Optional[float] = None
#     add_slow_label:    bool  = True
#     api_key:           str   = ""
#     marker_overrides:  Dict[int, dict] = field(default_factory=dict)


# # ─────────────────────────────────────────────────────────────────
# # Generated polygon
# # ─────────────────────────────────────────────────────────────────

# @dataclass
# class GenPoly:
#     marker_idx:    int
#     marker_name:   str
#     strip_idx:     int
#     coords:        List[Tuple[float, float]]
#     road_heading:  float
#     heading_src:   str
#     road_width_m:  float
#     lane_width_m:  float
#     strip_thick_m: float
#     strip_gap_m:   float
#     num_lanes:     int
#     along_offset_m: float
#     snapped_lat:   float
#     snapped_lon:   float


# # ─────────────────────────────────────────────────────────────────
# # Main generator
# # ─────────────────────────────────────────────────────────────────

# def generate_polygons(
#     markers: List[KMLMarker],
#     spec: PolySpec,
#     snapped: Optional[Dict[int, Tuple[float, float]]] = None,
# ) -> Tuple[List[GenPoly], Dict[int, Tuple[float, float, float, str]]]:
#     all_polys: List[GenPoly] = []
#     headings: Dict[int, Tuple[float, float, float, str]] = {}

#     for mk in markers:
#         i = mk.index
#         ov = spec.marker_overrides.get(i, {})

#         rw  = float(ov.get("road_width_m",      spec.road_width_m))
#         nl  = int(ov.get("num_lanes",            spec.num_lanes))
#         sep = float(ov.get("separator_width_m",  spec.separator_width_m))
#         ns  = int(ov.get("num_strips",           spec.num_strips))
#         st  = float(ov.get("strip_thick_m",      spec.strip_thick_m))
#         gap = float(ov.get("strip_gap_m",        spec.strip_gap_m))
#         lw  = (rw - sep) / max(nl, 1)

#         snap_lat = mk.lat
#         snap_lon = mk.lon
#         if snapped and i in snapped:
#             snap_lat, snap_lon = snapped[i]

#         manual_h = ov.get("heading_deg", spec.heading_override)
#         if manual_h is not None:
#             heading = norm180(float(manual_h))
#             hsrc    = "manual"
#         else:
#             heading, hsrc = detect_heading(markers, i, spec.api_key, snapped)

#         headings[i] = (heading, snap_lat, snap_lon, hsrc)

#         total_span = ns * st + (ns - 1) * gap
#         first_pos  = -total_span / 2 + st / 2

#         for si in range(ns):
#             along = first_pos + si * (st + gap)
#             coords = build_strip_rect(snap_lat, snap_lon, heading, along, rw, st)
#             all_polys.append(GenPoly(
#                 marker_idx     = i,
#                 marker_name    = mk.name,
#                 strip_idx      = si,
#                 coords         = coords,
#                 road_heading   = heading,
#                 heading_src    = hsrc,
#                 road_width_m   = rw,
#                 lane_width_m   = lw,
#                 strip_thick_m  = st,
#                 strip_gap_m    = gap,
#                 num_lanes      = nl,
#                 along_offset_m = along,
#                 snapped_lat    = snap_lat,
#                 snapped_lon    = snap_lon,
#             ))

#     return all_polys, headings


# # ─────────────────────────────────────────────────────────────────
# # Pipeline
# # ─────────────────────────────────────────────────────────────────

# def run_pipeline(
#     kml_path: str,
#     spec: PolySpec,
#     per_headings: Optional[Dict[int, float]] = None,
#     progress_cb=None,
# ) -> Tuple[List[KMLMarker], List[GenPoly], Dict]:
#     def _cb(msg, pct):
#         if progress_cb: progress_cb(msg, pct)

#     _cb("Parsing KML…", 10)
#     markers = parse_kml(kml_path)
#     if not markers:
#         raise ValueError("No Point markers found in KML.")

#     if per_headings:
#         for idx, hdg in per_headings.items():
#             spec.marker_overrides.setdefault(idx, {})["heading_deg"] = hdg

#     _cb("Snapping to roads via Google API…", 30)
#     snapped: Dict[int, Tuple[float, float]] = {}
#     if spec.api_key:
#         snapped = snap_to_roads(markers, spec.api_key)

#     _cb("Detecting road headings…", 50)
#     all_polys, headings = generate_polygons(markers, spec, snapped)

#     _cb("Done", 100)
#     return markers, all_polys, headings


# # ─────────────────────────────────────────────────────────────────
# # KML export
# # ─────────────────────────────────────────────────────────────────

# def _slow_label_kml(
#     snap_lat: float, snap_lon: float,
#     heading: float,
#     road_width_m: float,
#     strip_thick_m: float,
#     ahead_offset_m: float = 2.0,
# ) -> str:
#     label_lat, label_lon = offset_ll(snap_lat, snap_lon, heading, ahead_offset_m + strip_thick_m)
#     return (
#         f'<Placemark>'
#         f'<name>SLOW</name>'
#         f'<description>Speed breaker ahead warning</description>'
#         f'<styleUrl>#slow_label</styleUrl>'
#         f'<Point><coordinates>{label_lon:.8f},{label_lat:.8f},0</coordinates></Point>'
#         f'</Placemark>'
#     )


# def export_kml(
#     markers:   List[KMLMarker],
#     all_polys: List[GenPoly],
#     headings:  Dict,
#     spec:      PolySpec,
#     out_path:  str,
# ) -> None:
#     def cs(coords):
#         return " ".join(f"{lo:.8f},{la:.8f},0" for la, lo in coords)

#     by: Dict[int, List[GenPoly]] = {}
#     for p in all_polys: by.setdefault(p.marker_idx, []).append(p)

#     lines = [
#         '<?xml version="1.0" encoding="UTF-8"?>',
#         '<kml xmlns="http://www.opengis.net/kml/2.2">',
#         '<Document>',
#         '<name>CAP PTBM Speed Breaker Polygons — GIS BOQ v2 (p3)</name>',
#         # Yellow fill polygon
#         '<Style id="strip_yellow">',
#         '  <LineStyle><color>ff1400ff</color><width>2</width></LineStyle>',
#         '  <PolyStyle><color>cc00d7ff</color><fill>1</fill><outline>1</outline></PolyStyle>',
#         '</Style>',
#         # Orange fill for alternate strips
#         '<Style id="strip_orange">',
#         '  <LineStyle><color>ff1400ff</color><width>2</width></LineStyle>',
#         '  <PolyStyle><color>cc0088ff</color><fill>1</fill><outline>1</outline></PolyStyle>',
#         '</Style>',
#         # Red pushpin for center marker
#         '<Style id="pin">',
#         '  <IconStyle><scale>1.1</scale>',
#         '    <Icon><href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href></Icon>',
#         '  </IconStyle>',
#         '  <LabelStyle><scale>0.9</scale></LabelStyle>',
#         '</Style>',
#         # SLOW label style
#         '<Style id="slow_label">',
#         '  <IconStyle><scale>0</scale>',
#         '    <Icon><href>http://maps.google.com/mapfiles/kml/shapes/arrow.png</href></Icon>',
#         '  </IconStyle>',
#         '  <LabelStyle><color>ff00ffff</color><scale>1.4</scale></LabelStyle>',
#         '</Style>',
#     ]

#     for mk in markers:
#         ps = by.get(mk.index, [])
#         h_data = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
#         heading  = h_data[0]
#         snap_lat = h_data[1]
#         snap_lon = h_data[2]
#         hsrc     = h_data[3]
#         rw = ps[0].road_width_m  if ps else spec.road_width_m
#         nl = ps[0].num_lanes     if ps else spec.num_lanes
#         st = ps[0].strip_thick_m if ps else spec.strip_thick_m
#         sw_mm_val = st * 1000

#         lines.append(f'<Folder><name>{mk.name}</name>')

#         lines.append(
#             f'<Placemark><name>{mk.name}</name>'
#             f'<description>CAP PTBM {sw_mm_val:.0f}MM X {len(ps)}'
#             f'&#10;Road: {rw:.1f}m | {nl} Lane'
#             f'&#10;Heading: {heading:.1f}° [{hsrc}]'
#             f'&#10;Lat: {snap_lat:.6f} Lon: {snap_lon:.6f}</description>'
#             f'<styleUrl>#pin</styleUrl>'
#             f'<Point><coordinates>{snap_lon:.8f},{snap_lat:.8f},0</coordinates></Point>'
#             f'</Placemark>'
#         )

#         for p in ps:
#             style_id = "strip_yellow" if p.strip_idx % 2 == 0 else "strip_orange"
#             label = f"CAP PTBM {p.strip_thick_m*1000:.0f}MM X {len(ps)}"
#             lines.append(
#                 f'<Placemark><name>{label}</name>'
#                 f'<description>{mk.name} Strip {p.strip_idx+1}/{len(ps)}'
#                 f'&#10;Width (road): {p.road_width_m:.2f}m'
#                 f'&#10;Thick (along road): {p.strip_thick_m*1000:.0f}mm'
#                 f'&#10;Heading: {p.road_heading:.1f}°</description>'
#                 f'<styleUrl>#{style_id}</styleUrl>'
#                 f'<Polygon><outerBoundaryIs><LinearRing>'
#                 f'<coordinates>{cs(p.coords)}</coordinates>'
#                 f'</LinearRing></outerBoundaryIs></Polygon>'
#                 f'</Placemark>'
#             )

#         if spec.add_slow_label and ps:
#             lines.append(_slow_label_kml(snap_lat, snap_lon, heading, rw, st))

#         lines.append('</Folder>')

#     lines += ['</Document>', '</kml>']
#     with open(out_path, "w", encoding="utf-8") as f:
#         f.write("\n".join(lines))


# # ─────────────────────────────────────────────────────────────────
# # Excel export with embedded satellite images (high zoom)
# # ─────────────────────────────────────────────────────────────────

# def _tb() -> Border:
#     s = Side(style="thin")
#     return Border(left=s, right=s, top=s, bottom=s)

# def _fill(h: str) -> PatternFill:
#     return PatternFill("solid", start_color=h, end_color=h)

# def _hdr(ws, r, c, v, bg="1F3864", fg="FFFFFF"):
#     x = ws.cell(r, c, v)
#     x.font      = Font(bold=True, color=fg, size=9)
#     x.fill      = _fill(bg)
#     x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     x.border    = _tb()
#     return x

# def _dat(ws, r, c, v, bg=None, bold=False):
#     x = ws.cell(r, c, v)
#     x.font      = Font(bold=bold, size=9)
#     x.alignment = Alignment(horizontal="center", vertical="center")
#     x.border    = _tb()
#     if bg: x.fill = _fill(bg)
#     return x


# def export_excel(
#     markers:   List[KMLMarker],
#     all_polys: List[GenPoly],
#     headings:  Dict,
#     spec:      PolySpec,
#     out_path:  str,
#     progress_cb=None,
# ) -> None:
#     def _cb(msg, pct):
#         if progress_cb: progress_cb(msg, pct)

#     wb = openpyxl.Workbook()

#     # ── Sheet 1: BOQ Summary ─────────────────────────────────────
#     ws1 = wb.active
#     ws1.title = "BOQ Summary"

#     H1 = [
#         "S.No", "Placemark Name", "Latitude", "Longitude",
#         "Bearing (°)", "Road Width (m)", "Mark Length (m)",
#         "No. of Strips", "Area (m²)", "Rate (Rs/m²)", "Amount (Rs)",
#         "Satellite View",
#     ]
#     for ci, h in enumerate(H1, 1):
#         _hdr(ws1, 1, ci, h)
#     ws1.row_dimensions[1].height = 36

#     by: Dict[int, List[GenPoly]] = {}
#     for p in all_polys: by.setdefault(p.marker_idx, []).append(p)

#     # Use zoom=20 and scale=2 for high-resolution satellite image
#     IMAGE_ROW_H = 220

#     for ri, mk in enumerate(markers, 2):
#         ps    = by.get(mk.index, [])
#         p0    = ps[0] if ps else None
#         hdata = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
#         h_deg    = hdata[0]
#         snap_lat = hdata[1]
#         snap_lon = hdata[2]
#         rw    = p0.road_width_m  if p0 else spec.road_width_m
#         st    = p0.strip_thick_m if p0 else spec.strip_thick_m
#         gap   = p0.strip_gap_m   if p0 else spec.strip_gap_m
#         ns    = len(ps)
#         mark_len = ns * st + max(ns-1, 0) * gap
#         area_tot = round(st * rw * ns, 4)
#         rate     = 2500
#         amount   = round(area_tot * rate, 2)

#         bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
#         row_data = [
#             ri - 1, mk.name,
#             round(snap_lat, 6), round(snap_lon, 6),
#             round(h_deg, 1), round(rw, 2),
#             round(mark_len, 3), ns,
#             round(area_tot, 4), rate, amount, "",
#         ]
#         for ci, val in enumerate(row_data, 1):
#             _dat(ws1, ri, ci, val, bg=bg)

#         ws1.row_dimensions[ri].height = IMAGE_ROW_H

#         # Fetch high-zoom satellite image
#         if spec.api_key:
#             _cb(f"Fetching satellite image for {mk.name}…", 70)
#             # zoom=20 with scale=2 gives a very zoomed-in, high-res view
#             img_bytes = get_satellite_image(
#                 snap_lat, snap_lon, spec.api_key,
#                 zoom=20, size="640x480",
#             )
#             if img_bytes:
#                 try:
#                     img_stream = io.BytesIO(img_bytes)
#                     xl_img = XLImage(img_stream)
#                     xl_img.width  = 320
#                     xl_img.height = 210
#                     ws1.add_image(xl_img, f"L{ri}")
#                 except Exception:
#                     ws1.cell(ri, 12, "Image error")
#             else:
#                 ws1.cell(ri, 12, "Image unavailable (check API key/quota)")
#         else:
#             ws1.cell(ri, 12, "Add GOOGLE_API_KEY for satellite images")

#     for i, w in enumerate([5, 24, 12, 12, 10, 13, 14, 10, 12, 12, 14, 52], 1):
#         ws1.column_dimensions[get_column_letter(i)].width = w
#     ws1.column_dimensions["L"].width = 52

#     # ── Sheet 2: Strip Coordinates ───────────────────────────────
#     ws2 = wb.create_sheet("Strip Coordinates")
#     H2 = ["S.No", "Marker", "Strip #", "Along Offset (m)",
#           "C1 Lat", "C1 Lon", "C2 Lat", "C2 Lon",
#           "C3 Lat", "C3 Lon", "C4 Lat", "C4 Lon",
#           "Road Width (m)", "Thick (mm)", "Gap (m)", "Heading °"]
#     for ci, h in enumerate(H2, 1): _hdr(ws2, 1, ci, h, bg="145A32")
#     ws2.row_dimensions[1].height = 36

#     for gi, p in enumerate(all_polys, 1):
#         ri = gi + 1
#         bg = "E9F7EF" if ri % 2 == 0 else "FFFFFF"
#         cs_ = p.coords[:4]
#         row2 = [gi, p.marker_name, p.strip_idx+1, round(p.along_offset_m, 4),
#                 *[round(v, 8) for c_ in cs_ for v in c_],
#                 round(p.road_width_m, 3), round(p.strip_thick_m*1000, 1),
#                 round(p.strip_gap_m, 3), round(p.road_heading, 1)]
#         for ci, val in enumerate(row2, 1): _dat(ws2, ri, ci, val, bg=bg)
#     for i, w in enumerate([5, 22, 8, 14, 12, 12, 12, 12, 12, 12, 12, 12, 12, 10, 9, 10], 1):
#         ws2.column_dimensions[get_column_letter(i)].width = w

#     # ── Sheet 3: Spec Summary ────────────────────────────────────
#     ws3 = wb.create_sheet("Spec Summary")
#     _hdr(ws3, 1, 1, "Parameter", bg="784212")
#     _hdr(ws3, 1, 2, "Value", bg="784212")
#     ws3.column_dimensions["A"].width = 32
#     ws3.column_dimensions["B"].width = 60

#     # Count markers (need to pass it somehow — use all_polys)
#     n_markers_count = len(set(p.marker_idx for p in all_polys))

#     spec_rows = [
#         ("Tool",            "GIS BOQ Speed Breaker Tool v2 (p3)"),
#         ("Institute",       "IIIT Nagpur — Dr. Neha Kasture"),
#         ("Client",          "PWD / NHAI"),
#         ("", ""),
#         ("Strip Type",      "CAP PTBM (Capsule Prefab Thermoplastic Bituminous Marking)"),
#         ("Strip Orientation","PERPENDICULAR to road — spans full road width"),
#         ("Strip Thickness", f"{spec.strip_thick_m*1000:.0f} mm thick (along road) = {spec.strip_thick_m:.2f}m"),
#         ("Road Width",      f"{spec.road_width_m:.1f} m (strip length across road)"),
#         ("Number of Strips",str(spec.num_strips)),
#         ("Gap Between",     f"{spec.strip_gap_m:.2f} m (along road)"),
#         ("SLOW Label",      "Yes — added to KML 2m ahead of strips" if spec.add_slow_label else "No"),
#         ("Google API",      "Connected" if spec.api_key else "Not set — using neighbour heading"),
#         ("Heading Method",  "Google Roads API + snap-to-road" if spec.api_key else "Neighbour bearing average"),
#         ("Satellite Zoom",  "20 (max zoom, high-res 640x480 @2x scale)"),
#         ("", ""),
#         ("Total Markers",   str(n_markers_count)),
#         ("Total Strips",    str(len(all_polys))),
#         ("Rate (default)",  "Rs 2,500 / m²"),
#         ("", ""),
#         ("NOTE", "All markers assumed to be at CENTER of road. Polygon spans ±(road_width/2) cross-road."),
#     ]
#     for ri2, (k, v) in enumerate(spec_rows, 2):
#         ws3.cell(ri2, 1, k).font = Font(bold=bool(k), size=9)
#         ws3.cell(ri2, 2, v).font = Font(size=9)

#     wb.save(out_path)


# # ─────────────────────────────────────────────────────────────────
# # Self-test
# # ─────────────────────────────────────────────────────────────────
# if __name__ == "__main__":
#     TEST_KML = """<?xml version="1.0" encoding="UTF-8"?>
# <kml xmlns="http://www.opengis.net/kml/2.2"><Document>
#   <Placemark><name>SB_1</name><Point><coordinates>93.94313688,24.83678499,0</coordinates></Point></Placemark>
#   <Placemark><name>SB_2</name><Point><coordinates>93.94334266,24.83687056,0</coordinates></Point></Placemark>
#   <Placemark><name>SB_3</name><Point><coordinates>93.94342449,24.83672434,0</coordinates></Point></Placemark>
# </Document></kml>"""
#     with tempfile.NamedTemporaryFile(mode="w", suffix=".kml", delete=False) as f:
#         f.write(TEST_KML); kp = f.name

#     api_key = ""
#     env_path = r"D:\Road_Safety_Reasearch_III\.env"
#     if os.path.exists(env_path):
#         for line in open(env_path):
#             if line.startswith("GOOGLE_API_KEY="):
#                 api_key = line.split("=", 1)[1].strip()
#                 break

#     spec = PolySpec(
#         road_width_m=7.0, num_lanes=2, separator_width_m=0.5,
#         num_strips=3, strip_thick_m=0.5, strip_gap_m=0.6,
#         add_slow_label=True, api_key=api_key,
#     )
#     markers, polys, headings = run_pipeline(kp, spec)
#     print(f"\n✅ Markers: {len(markers)} | Strips: {len(polys)}")
#     for mk in markers:
#         ps = [p for p in polys if p.marker_idx == mk.index]
#         hd = headings[mk.index]
#         print(f"  📍 {mk.name}  snap=({hd[1]:.6f},{hd[2]:.6f})  "
#               f"heading={hd[0]:.1f}° [{hd[3]}]  strips={len(ps)}")

#     export_kml(markers, polys, headings, spec, "/tmp/p3_test.kml")
#     export_excel(markers, polys, headings, spec, "/tmp/p3_test.xlsx")
#     print("\n✅ KML + Excel exported to /tmp/")
#     os.unlink(kp)







"""
p3.py v5 — Speed Breaker CAP PTBM Polygon Engine
IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI

v5 FIXES:
  ● LABEL_PRESETS now properly exported at module level (fixes ImportError)
  ● 8 parallel workers for image fetching (was 4)
  ● Parallel heading detection via ThreadPoolExecutor
  ● Parallel snap-to-roads chunking
  ● Google Static Maps API (1 call/marker) + Pillow polygon overlay
  ● 7 real-life road labels, rename placemarks, curve-aware polygons
"""
from __future__ import annotations
import math, os, io, xml.etree.ElementTree as ET, tempfile
import concurrent.futures
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

R_EARTH = 6_371_000.0

# ── Exported constants (imported by ui3.py) ───────────────────────

LABEL_PRESETS: Dict[str, dict] = {
    "SLOW":          {"offset_m": 2.0, "scale": 1.5, "kml_color": "ff00ffff", "desc": "General speed warning"},
    "SPEED BREAKER": {"offset_m": 3.0, "scale": 1.3, "kml_color": "ff00aaff", "desc": "Speed breaker ahead"},
    "CAUTION":       {"offset_m": 2.0, "scale": 1.3, "kml_color": "ff0088ff", "desc": "General hazard warning"},
    "SCHOOL ZONE":   {"offset_m": 3.0, "scale": 1.2, "kml_color": "ff00cc44", "desc": "School zone — reduce speed"},
    "ROAD HUMP":     {"offset_m": 2.0, "scale": 1.2, "kml_color": "ffff8800", "desc": "Road hump / speed hump"},
    "RUMBLE STRIP":  {"offset_m": 2.0, "scale": 1.2, "kml_color": "ffff00ff", "desc": "Rumble strip marking"},
    "STOP":          {"offset_m": 2.5, "scale": 1.5, "kml_color": "ff0000ff", "desc": "Stop line marking"},
}

LANE_PRESETS: Dict[str, dict] = {
    "1-Lane (3.5m)":  dict(num_lanes=1, road_width_m=3.5,  separator_width_m=0.0),
    "2-Lane (7.0m)":  dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
    "4-Lane (14.0m)": dict(num_lanes=4, road_width_m=14.0, separator_width_m=2.0),
    "6-Lane (21.0m)": dict(num_lanes=6, road_width_m=21.0, separator_width_m=3.0),
    "Custom":         dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
}

# ─────────────────────────────────────────────────────────────────
# Geometry helpers
# ─────────────────────────────────────────────────────────────────

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1); dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R_EARTH * math.asin(math.sqrt(max(0.0, min(1.0, a))))

def forward_bearing(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dl = math.radians(lon2 - lon1)
    x = math.sin(dl) * math.cos(p2)
    y = math.cos(p1)*math.sin(p2) - math.sin(p1)*math.cos(p2)*math.cos(dl)
    return (math.degrees(math.atan2(x, y)) + 360) % 360

def norm180(h: float) -> float:
    h = h % 360; return h - 180 if h >= 180 else h

def offset_ll(lat: float, lon: float, bearing_deg: float, dist_m: float) -> Tuple[float, float]:
    d = dist_m / R_EARTH; b = math.radians(bearing_deg)
    p1 = math.radians(lat); l1 = math.radians(lon)
    p2 = math.asin(math.sin(p1)*math.cos(d) + math.cos(p1)*math.sin(d)*math.cos(b))
    l2 = l1 + math.atan2(math.sin(b)*math.sin(d)*math.cos(p1),
                          math.cos(d) - math.sin(p1)*math.sin(p2))
    return math.degrees(p2), math.degrees(l2)

def build_strip_rect(centre_lat: float, centre_lon: float, road_heading: float,
                     along_offset_m: float, road_width_m: float,
                     strip_thick_m: float) -> List[Tuple[float, float]]:
    sc_lat, sc_lon = offset_ll(centre_lat, centre_lon, road_heading, along_offset_m)
    perp = (road_heading + 90) % 360
    hw, ht = road_width_m / 2, strip_thick_m / 2
    ll = offset_ll(sc_lat, sc_lon, perp, +hw)
    rl = offset_ll(sc_lat, sc_lon, perp, -hw)
    lf = offset_ll(ll[0], ll[1], road_heading, +ht)
    lb = offset_ll(ll[0], ll[1], road_heading, -ht)
    rb = offset_ll(rl[0], rl[1], road_heading, -ht)
    rf = offset_ll(rl[0], rl[1], road_heading, +ht)
    return [lf, lb, rb, rf, lf]


# ─────────────────────────────────────────────────────────────────
# Google Static Maps + Pillow polygon overlay (FAST)
# ─────────────────────────────────────────────────────────────────

def _mercator_y(lat_deg: float) -> float:
    lat_r = math.radians(lat_deg)
    return (1.0 - math.log(math.tan(lat_r) + 1.0 / math.cos(lat_r)) / math.pi) / 2.0

def _ll_to_img_px(lat: float, lon: float,
                  center_lat: float, center_lon: float,
                  zoom: int, img_w: int, img_h: int,
                  scale: int = 2) -> Tuple[float, float]:
    """Web Mercator lat/lon → pixel in Static Maps image."""
    tile_px = 256 * scale
    n = 2.0 ** zoom

    def world(lt, lg):
        wx = (lg + 180.0) / 360.0 * n * tile_px
        wy = _mercator_y(lt) * n * tile_px
        return wx, wy

    cx_w, cy_w = world(center_lat, center_lon)
    px_w, py_w = world(lat, lon)
    return img_w / 2.0 + (px_w - cx_w), img_h / 2.0 + (py_w - cy_w)


def _image_is_dark(img_bytes: bytes, threshold: int = 30) -> bool:
    """Return True if average pixel brightness < threshold (black/missing tile)."""
    try:
        from PIL import Image as PILImage
        import struct
        img = PILImage.open(io.BytesIO(img_bytes)).convert("L")  # grayscale
        pixels = list(img.getdata())
        avg = sum(pixels) / max(len(pixels), 1)
        return avg < threshold
    except Exception:
        return False


def fetch_static_map(lat: float, lon: float, api_key: str,
                     zoom: int = 18, size: str = "640x640",
                     scale: int = 2) -> Tuple[Optional[bytes], int]:
    """
    Fetch Google Static Maps satellite image.
    Auto-falls back to lower zoom if image is dark/empty (no imagery at that zoom).
    Returns (image_bytes, actual_zoom_used).
    """
    url = "https://maps.googleapis.com/maps/api/staticmap"
    # Try requested zoom, then step down if dark
    for z in [zoom, zoom - 1, zoom - 2, 17, 16]:
        params = {
            "center":  f"{lat},{lon}",
            "zoom":    str(z),
            "size":    size,
            "maptype": "satellite",
            "scale":   str(scale),
            "key":     api_key,
        }
        try:
            r = requests.get(url, params=params, timeout=14)
            if r.status_code == 200 and r.headers.get("content-type", "").startswith("image"):
                if not _image_is_dark(r.content, threshold=25):
                    return r.content, z
                # Dark at this zoom — try lower
        except Exception:
            pass
    return None, zoom


def draw_polys_on_image(img_bytes: bytes,
                        polys: List["GenPoly"],
                        center_lat: float, center_lon: float,
                        zoom: int = 18, scale: int = 2) -> bytes:
    """
    Draw bright yellow polygon strips on satellite image.
    - Thick black outline so strips are visible on any background
    - Large center crosshair
    - Info bar at bottom
    """
    try:
        from PIL import Image as PILImage, ImageDraw, ImageFont
    except ImportError:
        return img_bytes

    img = PILImage.open(io.BytesIO(img_bytes)).convert("RGBA")
    w, h = img.size
    overlay = PILImage.new("RGBA", (w, h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)

    # Bright fills — high opacity so clearly visible
    FILLS = [
        (255, 230,  0, 230),   # bright yellow
        (255, 170,  0, 230),   # amber
        (255, 210,  0, 230),
        (200, 255,  0, 230),   # yellow-green
        (255, 140,  0, 230),
        (255, 255, 80, 230),
    ]
    OUTLINE_BLACK = (0,   0,   0, 255)   # thick black outer outline
    OUTLINE_WHITE = (255, 255, 255, 255) # thin white inner outline

    def px(lat, lon):
        return _ll_to_img_px(lat, lon, center_lat, center_lon, zoom, w, h, scale)

    for p in polys:
        pts = [px(la, lo) for la, lo in p.coords[:-1]]
        # 1. Black outer outline (thick, visible on any background)
        for i in range(len(pts)):
            draw.line([pts[i], pts[(i+1) % len(pts)]], fill=OUTLINE_BLACK, width=7)
        # 2. Filled polygon
        draw.polygon(pts, fill=FILLS[p.strip_idx % len(FILLS)])
        # 3. White inner outline
        for i in range(len(pts)):
            draw.line([pts[i], pts[(i+1) % len(pts)]], fill=OUTLINE_WHITE, width=2)

    # Crosshair at center marker
    cx, cy = px(center_lat, center_lon)
    arm = 18
    draw.line([(cx - arm, cy), (cx + arm, cy)], fill=(255, 50, 50, 255), width=4)
    draw.line([(cx, cy - arm), (cx, cy + arm)], fill=(255, 50, 50, 255), width=4)
    draw.ellipse([(cx-8, cy-8), (cx+8, cy+8)],
                 fill=(255, 30, 30, 230), outline=(255, 255, 255, 255))

    final = PILImage.alpha_composite(img, overlay).convert("RGB")

    # Info bar
    try:
        font = ImageFont.load_default()
        rw  = polys[0].road_width_m  if polys else 7.0
        thk = polys[0].strip_thick_m * 1000 if polys else 500
        ns  = len(polys)
        bd  = ImageDraw.Draw(final)
        bh  = 22
        bd.rectangle([(0, h - bh), (w, h)], fill=(0, 0, 0))
        bd.text((6, h - bh + 4),
                f"W={rw:.0f}m  T={thk:.0f}mm  Strips={ns}  Zoom={zoom}  Google Satellite",
                fill=(255, 215, 0), font=font)
    except Exception:
        pass

    buf = io.BytesIO()
    final.save(buf, format="PNG")
    return buf.getvalue()


def _esri_tile_fallback(snap_lat: float, snap_lon: float,
                        polys: List["GenPoly"],
                        zoom: int = 19,
                        img_w: int = 640, img_h: int = 640) -> Optional[bytes]:
    """ESRI tile stitching — used only when no API key."""
    try:
        from PIL import Image as PILImage, ImageDraw, ImageFont
    except ImportError:
        return None

    TILE_PX = 256

    def _ll_tile(lat, lon, z):
        n = 2**z
        tx = int((lon + 180.0) / 360.0 * n)
        lr = math.radians(lat)
        ty = int((1.0 - math.log(math.tan(lr) + 1.0/math.cos(lr))/math.pi)/2.0*n)
        return tx, ty

    def _ll_px(lat, lon, z, otx, oty):
        n = 2**z
        px = (lon + 180.0)/360.0*n*TILE_PX - otx*TILE_PX
        lr = math.radians(lat)
        py = (1.0 - math.log(math.tan(lr)+1.0/math.cos(lr))/math.pi)/2.0*n*TILE_PX - oty*TILE_PX
        return px, py

    cx_t, cy_t = _ll_tile(snap_lat, snap_lon, zoom)
    tiles_x = math.ceil(img_w/TILE_PX) + 3
    tiles_y = math.ceil(img_h/TILE_PX) + 3
    otx = cx_t - tiles_x//2; oty = cy_t - tiles_y//2
    canvas = PILImage.new("RGB", (tiles_x*TILE_PX, tiles_y*TILE_PX), (40, 40, 40))

    hdrs = {"User-Agent": "Mozilla/5.0 (GIS-BOQ/5)"}

    # Parallel tile fetch
    def fetch_tile(dx_dy):
        dx, dy = dx_dy
        tx, ty = otx+dx, oty+dy
        url = (f"https://server.arcgisonline.com/ArcGIS/rest/services/"
               f"World_Imagery/MapServer/tile/{zoom}/{ty}/{tx}")
        try:
            r = requests.get(url, headers=hdrs, timeout=8)
            if r.status_code == 200:
                return dx, dy, r.content
        except Exception:
            pass
        return dx, dy, None

    coords_list = [(dx, dy) for dx in range(tiles_x) for dy in range(tiles_y)]
    with concurrent.futures.ThreadPoolExecutor(max_workers=16) as pool:
        for dx, dy, content in pool.map(fetch_tile, coords_list):
            if content:
                try:
                    t = PILImage.open(io.BytesIO(content)).convert("RGB")
                    canvas.paste(t, (dx*TILE_PX, dy*TILE_PX))
                except Exception:
                    pass

    cx_px, cy_px = _ll_px(snap_lat, snap_lon, zoom, otx, oty)
    left = max(0, int(cx_px - img_w//2))
    top  = max(0, int(cy_px - img_h//2))
    cropped = canvas.crop((left, top, left+img_w, top+img_h)).convert("RGBA")

    overlay = PILImage.new("RGBA", (img_w, img_h), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    FILLS = [(255,230,0,230),(255,165,0,230),(255,200,0,230),(200,255,0,230),(255,140,0,230),(255,255,80,230)]

    def tp(lat, lon):
        px2, py2 = _ll_px(lat, lon, zoom, otx, oty)
        return px2-left, py2-top

    for p in polys:
        pts = [tp(la, lo) for la, lo in p.coords[:-1]]
        # Black outer outline, bright fill, white inner
        for i in range(len(pts)):
            draw.line([pts[i], pts[(i+1)%len(pts)]], fill=(0,0,0,255), width=7)
        draw.polygon(pts, fill=FILLS[p.strip_idx % len(FILLS)])
        for i in range(len(pts)):
            draw.line([pts[i], pts[(i+1)%len(pts)]], fill=(255,255,255,255), width=2)

    cx_i, cy_i = tp(snap_lat, snap_lon)
    arm = 18
    draw.line([(cx_i-arm,cy_i),(cx_i+arm,cy_i)], fill=(255,50,50,255), width=4)
    draw.line([(cx_i,cy_i-arm),(cx_i,cy_i+arm)], fill=(255,50,50,255), width=4)
    draw.ellipse([(cx_i-8,cy_i-8),(cx_i+8,cy_i+8)],
                 fill=(255,30,30,220), outline=(255,255,255,255))

    final = PILImage.alpha_composite(cropped, overlay).convert("RGB")
    try:
        font = ImageFont.load_default()
        rw = polys[0].road_width_m if polys else 7.0
        thk = polys[0].strip_thick_m * 1000 if polys else 500
        d2 = ImageDraw.Draw(final)
        d2.rectangle([(0, img_h-22), (img_w, img_h)], fill=(0, 0, 0))
        d2.text((6, img_h-18),
                f"W={rw:.0f}m T={thk:.0f}mm Strips={len(polys)} ESRI Satellite",
                fill=(255, 215, 0), font=font)
    except Exception:
        pass

    buf = io.BytesIO()
    final.save(buf, format="PNG")
    return buf.getvalue()


def capture_polygon_image_fast(snap_lat: float, snap_lon: float,
                                polys: List["GenPoly"],
                                api_key: str,
                                zoom: int = 18) -> Optional[bytes]:
    """
    FAST: 1 Google Static Maps call + Pillow polygon overlay.
    Auto-falls back to lower zoom if image is dark/missing.
    Fallback to ESRI tiles if no API key.
    zoom=18: best balance of road detail + surrounding context for India.
    """
    if api_key:
        raw, actual_zoom = fetch_static_map(
            snap_lat, snap_lon, api_key,
            zoom=zoom, size="640x640", scale=2)
        if raw:
            return draw_polys_on_image(
                raw, polys, snap_lat, snap_lon, actual_zoom, scale=2)
    # Fallback: ESRI tiles
    for z in [18, 17]:
        result = _esri_tile_fallback(snap_lat, snap_lon, polys, z)
        if result: return result
    return None


# ─────────────────────────────────────────────────────────────────
# Google Maps Roads API (snap + heading)
# ─────────────────────────────────────────────────────────────────

def snap_to_roads(markers: List["KMLMarker"],
                  api_key: str) -> Dict[int, Tuple[float, float]]:
    if not api_key or not markers:
        return {}
    snapped: Dict[int, Tuple[float, float]] = {}

    def _snap_chunk(chunk):
        path = "|".join(f"{m.lat},{m.lon}" for m in chunk)
        try:
            resp = requests.get(
                "https://roads.googleapis.com/v1/snapToRoads",
                params={"path": path, "interpolate": "false", "key": api_key},
                timeout=12)
            data = resp.json()
            result = {}
            if "snappedPoints" in data:
                for pt in data["snappedPoints"]:
                    oi = pt.get("originalIndex", -1)
                    if oi >= 0:
                        loc = pt["location"]
                        result[chunk[oi].index] = (loc["latitude"], loc["longitude"])
            return result
        except Exception:
            return {}

    chunk_size = 100
    chunks = [markers[i:i+chunk_size] for i in range(0, len(markers), chunk_size)]
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as pool:
        for partial in pool.map(_snap_chunk, chunks):
            snapped.update(partial)
    return snapped


def get_road_heading_from_maps(lat: float, lon: float,
                                api_key: str,
                                radius_m: float = 30) -> Optional[float]:
    if not api_key:
        return None
    o1 = offset_ll(lat, lon, 0, radius_m)
    o2 = offset_ll(lat, lon, 90, radius_m)
    path = f"{lat},{lon}|{o1[0]},{o1[1]}|{o2[0]},{o2[1]}"
    try:
        resp = requests.get(
            "https://roads.googleapis.com/v1/snapToRoads",
            params={"path": path, "interpolate": "true", "key": api_key},
            timeout=8)
        pts = resp.json().get("snappedPoints", [])
        if len(pts) >= 2:
            p1, p2 = pts[0]["location"], pts[1]["location"]
            return norm180(forward_bearing(
                p1["latitude"], p1["longitude"],
                p2["latitude"], p2["longitude"]))
    except Exception:
        pass
    return None


# ─────────────────────────────────────────────────────────────────
# KML parsing
# ─────────────────────────────────────────────────────────────────

@dataclass
class KMLMarker:
    name:  str
    lat:   float
    lon:   float
    index: int = 0


def parse_kml(path: str) -> List[KMLMarker]:
    root = ET.parse(path).getroot()
    markers: List[KMLMarker] = []

    def iter_pm(node):
        for c in node:
            if c.tag.split("}")[-1] == "Placemark": yield c
            else: yield from iter_pm(c)

    idx = 0
    for pm in iter_pm(root):
        name = f"Marker_{idx+1}"
        for c in pm:
            if c.tag.split("}")[-1] == "name":
                name = (c.text or "").strip() or name; break
        for el in pm.iter():
            if el.tag.split("}")[-1] == "Point":
                for ce in el.iter():
                    if ce.tag.split("}")[-1] == "coordinates" and ce.text:
                        p = ce.text.strip().split(",")
                        if len(p) >= 2:
                            try:
                                markers.append(
                                    KMLMarker(name, float(p[1]), float(p[0]), idx))
                                idx += 1
                            except ValueError:
                                pass
                break
    return markers


# ─────────────────────────────────────────────────────────────────
# Heading detection  (single marker — called in parallel)
# ─────────────────────────────────────────────────────────────────

def detect_heading(markers: List[KMLMarker], idx: int,
                   api_key: str = "",
                   snapped: Optional[Dict[int, Tuple[float, float]]] = None,
                   window: int = 3) -> Tuple[float, str]:
    mk = markers[idx]
    lat, lon = mk.lat, mk.lon
    if snapped and idx in snapped:
        lat, lon = snapped[idx]

    if api_key:
        h = get_road_heading_from_maps(lat, lon, api_key)
        if h is not None:
            return h, "Google Roads API"

    bearings: List[Tuple[float, float]] = []
    for off in range(-window, window + 1):
        if off == 0: continue
        j = idx + off
        if j < 0 or j >= len(markers): continue
        nb = markers[j]
        nb_lat, nb_lon = nb.lat, nb.lon
        if snapped and j in snapped:
            nb_lat, nb_lon = snapped[j]
        d = haversine(lat, lon, nb_lat, nb_lon)
        if d < 0.5: continue
        b = norm180(forward_bearing(lat, lon, nb_lat, nb_lon))
        w = 1.0 / (abs(off) * max(d, 1.0))
        bearings.append((b, w))

    if not bearings:
        return 0.0, "default (0°)"
    sx = sum(w * math.cos(math.radians(2*b)) for b, w in bearings)
    sy = sum(w * math.sin(math.radians(2*b)) for b, w in bearings)
    return norm180(math.degrees(math.atan2(sy, sx)) / 2), "neighbour-avg"


# ─────────────────────────────────────────────────────────────────
# Spec & GenPoly
# ─────────────────────────────────────────────────────────────────

@dataclass
class PolySpec:
    road_width_m:      float = 7.0
    num_lanes:         int   = 2
    separator_width_m: float = 0.5
    num_strips:        int   = 3
    strip_thick_m:     float = 0.5
    strip_gap_m:       float = 0.6
    heading_override:  Optional[float] = None
    labels:            List[str] = field(default_factory=lambda: ["SLOW"])
    rename_placemarks: bool  = True
    api_key:           str   = ""
    marker_overrides:  Dict[int, dict] = field(default_factory=dict)


@dataclass
class GenPoly:
    marker_idx:     int
    marker_name:    str
    strip_idx:      int
    coords:         List[Tuple[float, float]]
    road_heading:   float
    heading_src:    str
    road_width_m:   float
    lane_width_m:   float
    strip_thick_m:  float
    strip_gap_m:    float
    num_lanes:      int
    along_offset_m: float
    snapped_lat:    float
    snapped_lon:    float


# ─────────────────────────────────────────────────────────────────
# Polygon generator — headings detected in parallel
# ─────────────────────────────────────────────────────────────────

def generate_polygons(markers: List[KMLMarker], spec: PolySpec,
                      snapped: Optional[Dict[int, Tuple[float, float]]] = None
                      ) -> Tuple[List[GenPoly], Dict]:
    all_polys: List[GenPoly] = []
    headings: Dict[int, Tuple] = {}

    # Parallel heading detection via Google Roads API
    def _detect(mk_idx):
        mk = markers[mk_idx]
        ov = spec.marker_overrides.get(mk.index, {})
        manual_h = ov.get("heading_deg", spec.heading_override)
        if manual_h is not None:
            return mk.index, (norm180(float(manual_h)), mk.lat, mk.lon, "manual")
        h, hsrc = detect_heading(markers, mk_idx, spec.api_key, snapped)
        snap_lat, snap_lon = mk.lat, mk.lon
        if snapped and mk.index in snapped:
            snap_lat, snap_lon = snapped[mk.index]
        return mk.index, (h, snap_lat, snap_lon, hsrc)

    # Use 8 workers for parallel heading API calls
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        for mk_index, hdata in pool.map(_detect, range(len(markers))):
            headings[mk_index] = hdata

    for mk in markers:
        i = mk.index
        ov  = spec.marker_overrides.get(i, {})
        rw  = float(ov.get("road_width_m",     spec.road_width_m))
        nl  = int(ov.get("num_lanes",           spec.num_lanes))
        sep = float(ov.get("separator_width_m", spec.separator_width_m))
        ns  = int(ov.get("num_strips",          spec.num_strips))
        st  = float(ov.get("strip_thick_m",     spec.strip_thick_m))
        gap = float(ov.get("strip_gap_m",       spec.strip_gap_m))
        lw  = (rw - sep) / max(nl, 1)

        hdata = headings[i]
        heading   = hdata[0]
        snap_lat  = hdata[1]
        snap_lon  = hdata[2]

        total_span = ns * st + (ns - 1) * gap
        first_pos  = -total_span / 2 + st / 2

        for si in range(ns):
            along = first_pos + si * (st + gap)
            coords = build_strip_rect(snap_lat, snap_lon, heading, along, rw, st)
            all_polys.append(GenPoly(
                marker_idx=i, marker_name=mk.name, strip_idx=si,
                coords=coords, road_heading=heading, heading_src=hdata[3],
                road_width_m=rw, lane_width_m=lw, strip_thick_m=st,
                strip_gap_m=gap, num_lanes=nl, along_offset_m=along,
                snapped_lat=snap_lat, snapped_lon=snap_lon,
            ))

    return all_polys, headings


# ─────────────────────────────────────────────────────────────────
# Pipeline
# ─────────────────────────────────────────────────────────────────

def run_pipeline(kml_path: str, spec: PolySpec,
                 per_headings: Optional[Dict[int, float]] = None,
                 progress_cb=None):
    def _cb(msg, pct):
        if progress_cb: progress_cb(msg, pct)

    _cb("Parsing KML…", 10)
    markers = parse_kml(kml_path)
    if not markers:
        raise ValueError("No Point markers found in KML.")

    if per_headings:
        for idx, hdg in per_headings.items():
            spec.marker_overrides.setdefault(idx, {})["heading_deg"] = hdg

    _cb("Snapping to roads (parallel)…", 25)
    snapped: Dict[int, Tuple[float, float]] = {}
    if spec.api_key:
        snapped = snap_to_roads(markers, spec.api_key)

    _cb("Detecting headings (parallel)…", 45)
    all_polys, headings = generate_polygons(markers, spec, snapped)
    _cb("Done", 100)
    return markers, all_polys, headings


# ─────────────────────────────────────────────────────────────────
# KML export
# ─────────────────────────────────────────────────────────────────

def _label_styles_kml(labels: List[str]) -> List[str]:
    lines = []
    for label in labels:
        lp = LABEL_PRESETS.get(label, LABEL_PRESETS["SLOW"])
        sid = f"lbl_{label.replace(' ', '_')}"
        lines += [
            f'<Style id="{sid}">',
            f'  <IconStyle><scale>0</scale>',
            f'    <Icon><href>http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png</href></Icon>',
            f'  </IconStyle>',
            f'  <LabelStyle><color>{lp["kml_color"]}</color><scale>{lp["scale"]}</scale></LabelStyle>',
            f'</Style>',
        ]
    return lines


def _label_placemark(snap_lat: float, snap_lon: float, heading: float,
                     label: str, label_idx: int, strip_thick_m: float) -> str:
    lp = LABEL_PRESETS.get(label, LABEL_PRESETS["SLOW"])
    offset_m = lp["offset_m"] + label_idx * 3.0 + strip_thick_m
    ll_lat, ll_lon = offset_ll(snap_lat, snap_lon, heading, offset_m)
    sid = f"lbl_{label.replace(' ', '_')}"
    return (f'<Placemark><n>{label}</n>'
            f'<description>{lp["desc"]}</description>'
            f'<styleUrl>#{sid}</styleUrl>'
            f'<Point><coordinates>{ll_lon:.8f},{ll_lat:.8f},0</coordinates></Point>'
            f'</Placemark>')


def export_kml(markers: List[KMLMarker], all_polys: List[GenPoly],
               headings: Dict, spec: PolySpec, out_path: str) -> None:
    def cs(coords):
        return " ".join(f"{lo:.8f},{la:.8f},0" for la, lo in coords)

    by: dict = {}
    for p in all_polys: by.setdefault(p.marker_idx, []).append(p)
    labels = spec.labels if spec.labels else ["SLOW"]

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '<Document>',
        '<n>CAP PTBM Speed Breaker Polygons — GIS BOQ v5</n>',
        '<Style id="strip_a"><LineStyle><color>ff1400ff</color><width>2</width></LineStyle>'
        '<PolyStyle><color>cc00d7ff</color><fill>1</fill><outline>1</outline></PolyStyle></Style>',
        '<Style id="strip_b"><LineStyle><color>ff1400ff</color><width>2</width></LineStyle>'
        '<PolyStyle><color>cc0088ff</color><fill>1</fill><outline>1</outline></PolyStyle></Style>',
        '<Style id="pin"><IconStyle><scale>1.1</scale>'
        '<Icon><href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href></Icon>'
        '</IconStyle><LabelStyle><scale>0.9</scale></LabelStyle></Style>',
    ]
    lines.extend(_label_styles_kml(labels))

    for mk in markers:
        ps = by.get(mk.index, [])
        hd = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
        heading, snap_lat, snap_lon, hsrc = hd[0], hd[1], hd[2], hd[3]
        rw = ps[0].road_width_m  if ps else spec.road_width_m
        nl = ps[0].num_lanes     if ps else spec.num_lanes
        st = ps[0].strip_thick_m if ps else spec.strip_thick_m

        is_untitled = ("untitled" in mk.name.lower() or
                       mk.name.startswith("Marker_"))
        if spec.rename_placemarks and labels and is_untitled:
            display_name = f"{labels[0]} #{mk.index + 1}"
        else:
            display_name = mk.name

        lines.append(f'<Folder><n>{display_name}</n>')
        lines.append(
            f'<Placemark><n>{display_name}</n>'
            f'<description>CAP PTBM {st*1000:.0f}MM X {len(ps)}'
            f'&#10;Road: {rw:.1f}m | {nl} Lane | Heading: {heading:.1f}° [{hsrc}]'
            f'&#10;Lat: {snap_lat:.6f} Lon: {snap_lon:.6f}'
            f'&#10;Labels: {", ".join(labels)}</description>'
            f'<styleUrl>#pin</styleUrl>'
            f'<Point><coordinates>{snap_lon:.8f},{snap_lat:.8f},0</coordinates></Point>'
            f'</Placemark>'
        )
        for p in ps:
            sty = "strip_a" if p.strip_idx % 2 == 0 else "strip_b"
            lines.append(
                f'<Placemark><n>CAP PTBM {p.strip_thick_m*1000:.0f}MM X {len(ps)}</n>'
                f'<description>{display_name} Strip {p.strip_idx+1}/{len(ps)}'
                f'&#10;Width: {p.road_width_m:.2f}m | Thick: {p.strip_thick_m*1000:.0f}mm'
                f'&#10;Heading: {p.road_heading:.1f}°</description>'
                f'<styleUrl>#{sty}</styleUrl>'
                f'<Polygon><outerBoundaryIs><LinearRing>'
                f'<coordinates>{cs(p.coords)}</coordinates>'
                f'</LinearRing></outerBoundaryIs></Polygon>'
                f'</Placemark>'
            )
        for li, label in enumerate(labels):
            lines.append(_label_placemark(snap_lat, snap_lon, heading, label, li, st))
        lines.append('</Folder>')

    lines += ['</Document>', '</kml>']
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ─────────────────────────────────────────────────────────────────
# Excel export — 8 parallel workers
# ─────────────────────────────────────────────────────────────────

def _tb():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def _hdr(ws, r, c, v, bg="1F3864", fg="FFFFFF"):
    x = ws.cell(r, c, v)
    x.font = Font(bold=True, color=fg, size=9)
    x.fill = _fill(bg)
    x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    x.border = _tb()
    return x

def _dat(ws, r, c, v, bg=None):
    x = ws.cell(r, c, v)
    x.font = Font(size=9)
    x.alignment = Alignment(horizontal="center", vertical="center")
    x.border = _tb()
    if bg: x.fill = _fill(bg)
    return x


# def _render_worker(args):
#     """Called by each thread — returns (mk_index, png_bytes or None)."""
#     mk_index, snap_lat, snap_lon, ps, api_key = args
#     img_b = capture_polygon_image_fast(snap_lat, snap_lon, ps, api_key, zoom=18)
#     return mk_index, img_b

def _render_worker(args):
    """Called by each thread — returns (mk_index, png_bytes or None)."""
    mk_index, snap_lat, snap_lon, ps, api_key = args
    img_b = capture_polygon_image_fast(snap_lat, snap_lon, ps, api_key, zoom=18)
    return mk_index, img_b





def export_excel(markers: List[KMLMarker], all_polys: List[GenPoly],
                 headings: Dict, spec: PolySpec,
                 out_path: str, progress_cb=None) -> None:
    def _cb(msg, pct):
        if progress_cb: progress_cb(msg, pct)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "BOQ Summary"

    H1 = ["S.No", "Placemark Name", "Latitude", "Longitude",
          "Bearing (°)", "Road Width (m)", "Mark Length (m)",
          "No. of Strips", "Area (m²)", "Rate (Rs/m²)", "Amount (Rs)",
          "Satellite View  (Yellow = Polygon Strips)"]
    for ci, h in enumerate(H1, 1):
        _hdr(ws1, 1, ci, h)
    ws1.row_dimensions[1].height = 42

    by: dict = {}
    for p in all_polys: by.setdefault(p.marker_idx, []).append(p)

    # Write data rows first
    row_meta: dict = {}
    for ri, mk in enumerate(markers, 2):
        ps    = by.get(mk.index, [])
        p0    = ps[0] if ps else None
        hdata = headings.get(mk.index, (0.0, mk.lat, mk.lon, "default"))
        h_deg, snap_lat, snap_lon = hdata[0], hdata[1], hdata[2]
        rw   = p0.road_width_m  if p0 else spec.road_width_m
        st   = p0.strip_thick_m if p0 else spec.strip_thick_m
        gap  = p0.strip_gap_m   if p0 else spec.strip_gap_m
        ns   = len(ps)
        span = ns * st + max(ns - 1, 0) * gap
        area = round(st * rw * ns, 4)
        amt  = round(area * 2500, 2)

        bg = "F2F7FF" if ri % 2 == 0 else "FFFFFF"
        for ci, val in enumerate(
            [ri-1, mk.name, round(snap_lat,6), round(snap_lon,6),
             round(h_deg,1), round(rw,2), round(span,3), ns,
             round(area,4), 2500, amt, ""], 1
        ):
            _dat(ws1, ri, ci, val, bg=bg)
        ws1.row_dimensions[ri].height = 240
        row_meta[mk.index] = (ri, snap_lat, snap_lon, ps)

    # ── 8 parallel image renders ───────────────────────────────
    _cb(f"Fetching {len(markers)} satellite images (8 workers)…", 60)
    tasks = [
        (mk.index,
         row_meta[mk.index][1], row_meta[mk.index][2],
         row_meta[mk.index][3], spec.api_key)
        for mk in markers if mk.index in row_meta
    ]
    images: Dict[int, Optional[bytes]] = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_render_worker, t): t[0] for t in tasks}
        done = 0
        for fut in concurrent.futures.as_completed(futures):
            done += 1
            mk_idx, img_b = fut.result()
            images[mk_idx] = img_b
            pct = 60 + int(done / max(len(tasks), 1) * 30)
            _cb(f"Images: {done}/{len(tasks)} rendered…", pct)

    # Insert images
    for mk in markers:
        if mk.index not in row_meta: continue
        ri  = row_meta[mk.index][0]
        img_b = images.get(mk.index)
        if img_b:
            try:
                xl_img = XLImage(io.BytesIO(img_b))
                xl_img.width  = 360
                xl_img.height = 235
                ws1.add_image(xl_img, f"L{ri}")
            except Exception as e:
                ws1.cell(ri, 12, f"Image error: {e}")
        else:
            ws1.cell(ri, 12, "pip install Pillow — or check API key quota")

    for i, w in enumerate([5, 24, 12, 12, 10, 13, 14, 10, 12, 12, 14, 60], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.column_dimensions["L"].width = 60

    # ── Sheet 2: Strip Coordinates ────────────────────────────
    ws2 = wb.create_sheet("Strip Coordinates")
    H2 = ["S.No","Marker","Strip #","Along Offset (m)",
          "C1 Lat","C1 Lon","C2 Lat","C2 Lon",
          "C3 Lat","C3 Lon","C4 Lat","C4 Lon",
          "Road Width (m)","Thick (mm)","Gap (m)","Heading °"]
    for ci, h in enumerate(H2, 1): _hdr(ws2, 1, ci, h, bg="145A32")
    ws2.row_dimensions[1].height = 36
    for gi, p in enumerate(all_polys, 1):
        ri = gi + 1; bg = "E9F7EF" if ri % 2 == 0 else "FFFFFF"
        cs_ = p.coords[:4]
        row2 = [gi, p.marker_name, p.strip_idx+1, round(p.along_offset_m, 4),
                *[round(v, 8) for c_ in cs_ for v in c_],
                round(p.road_width_m,3), round(p.strip_thick_m*1000,1),
                round(p.strip_gap_m,3), round(p.road_heading,1)]
        for ci, val in enumerate(row2, 1): _dat(ws2, ri, ci, val, bg=bg)
    for i, w in enumerate([5,22,8,14,12,12,12,12,12,12,12,12,12,10,9,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Spec ─────────────────────────────────────────
    ws3 = wb.create_sheet("Spec Summary")
    _hdr(ws3, 1, 1, "Parameter", bg="784212")
    _hdr(ws3, 1, 2, "Value",     bg="784212")
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 90

    n_m = len(set(p.marker_idx for p in all_polys))
    spec_rows = [
        ("Tool",             "GIS BOQ Speed Breaker Tool v5"),
        ("Institute",        "IIIT Nagpur — Dr. Neha Kasture"),
        ("Client",           "PWD / NHAI"),
        ("",""),
        ("Strip Type",       "CAP PTBM — Capsule Prefab Thermoplastic Bituminous Marking"),
        ("Strip Orientation","PERPENDICULAR to road"),
        ("Strip Thickness",  f"{spec.strip_thick_m*1000:.0f} mm along road"),
        ("Road Width",       f"{spec.road_width_m:.1f} m (IRC carriageway)"),
        ("No. of Strips",    str(spec.num_strips)),
        ("Labels",           ", ".join(spec.labels)),
        ("Image Method",     "Google Static Maps (1 call/marker) + Pillow polygon overlay"),
        ("Workers",          "8 parallel threads for image rendering"),
        ("",""),
        ("Total Markers",    str(n_m)),
        ("Total Strips",     str(len(all_polys))),
        ("Rate",             "Rs 2,500 / m²"),
        ("",""),
        ("ROAD WIDTH NOTE",
         "Polygon is geodesically accurate. Visual narrowness in satellite is due to "
         "oblique satellite angle + shoulders outside IRC carriageway. "
         "Fix: measure with Google Earth Ruler → use Custom lane type."),
    ]
    for ri2, (k, v) in enumerate(spec_rows, 2):
        ws3.cell(ri2, 1, k).font = Font(bold=bool(k), size=9)
        c2 = ws3.cell(ri2, 2, v)
        c2.font = Font(size=9)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[ri2].height = 30 if v and len(v) > 80 else 15

    _cb("Saving Excel…", 96)
    wb.save(out_path)
# ─────────────────────────────────────────────────────────────────
# Self-test
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if any("streamlit" in arg for arg in sys.argv):
        pass  # Don't run self-test when launched via Streamlit
    else:
        TEST_KML = """<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2"><Document>
  <Placemark><n>SB_1</n><Point><coordinates>93.94313688,24.83678499,0</coordinates></Point></Placemark>
  <Placemark><n>SB_2</n><Point><coordinates>93.94334266,24.83687056,0</coordinates></Point></Placemark>
</Document></kml>"""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".kml", delete=False) as f:
            f.write(TEST_KML); kp = f.name

        api_key = ""
        for env_p in [r"D:\Road_Safety_Reasearch_III\.env", ".env"]:
            if os.path.exists(env_p):
                for line in open(env_p):
                    if line.startswith("GOOGLE_API_KEY="):
                        api_key = line.split("=", 1)[1].strip(); break

        spec = PolySpec(road_width_m=7.0, num_lanes=2, num_strips=3,
                        strip_thick_m=0.5, labels=["SLOW", "SPEED BREAKER"],
                        rename_placemarks=True, api_key=api_key)
        markers, polys, headings = run_pipeline(kp, spec)
        print(f"✅ Markers:{len(markers)} Strips:{len(polys)}")
        print(f"   LABEL_PRESETS keys: {list(LABEL_PRESETS.keys())}")
        export_kml(markers, polys, headings, spec, "/tmp/p3_v5.kml")
        export_excel(markers, polys, headings, spec, "/tmp/p3_v5.xlsx")
        print("✅ KML + Excel → /tmp/p3_v5.*")
        os.unlink(kp)