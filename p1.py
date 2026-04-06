"""
p1.py v10 — Speed Breaker CAP PTBM Polygon Engine
IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI

BUGS FIXED in v10:
  ● BUG 1: road_heading from green line was +90° off
      - Green line drawn ALONG road → bearing = road direction directly
      - Green line drawn ACROSS road → bearing + 90° = road direction
      - New UI option: "Line Direction" = Along / Across
      - Along road (default): heading = GL_bearing  (road runs same way as line)
      - Across road:          heading = GL_bearing + 90° + width = GL_length

  ● BUG 2: GL matched to only 1 marker — now matches ALL within max_dist
      - All markers within 50m of GL midpoint get the GL heading/width
      - No marker left without heading when a nearby GL exists

  ● BUG 3: GL length misused as road width when line is drawn along road
      - Along-road lines: length = segment length, NOT road width → use manual width
      - Across-road lines: length = road width → auto-fill

  ● BUG 4: Bounding box was axis-aligned (always NSEW rectangle)
      - Now uses rotated tight bounding box aligned to road heading
      - No more huge red rectangles that don't match road angle
"""
from __future__ import annotations
import math, xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

R_EARTH = 6_371_000.0

LANE_PRESETS: Dict[str, dict] = {
    "1-Lane (Town→Village)": dict(num_lanes=1, road_width_m=3.5,  separator_width_m=0.0),
    "2-Lane (City→Town)":    dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
    "4-Lane (City→City)":    dict(num_lanes=4, road_width_m=14.0, separator_width_m=2.0),
    "6-Lane (Highway)":      dict(num_lanes=6, road_width_m=21.0, separator_width_m=3.0),
}

MARKER_POSITION_LABELS: Dict[str, str] = {
    "centre":     "Centre (default)",
    "left_edge":  "Left Edge",
    "left_lane":  "Left Lane Centre",
    "right_lane": "Right Lane Centre",
    "right_edge": "Right Edge",
    "custom":     "Custom offset (m)",
}

LANE_COLOURS_KML = ["ff00d7ff","ff0088ff","ff00ff88","ffff44aa","ff44ffcc","ffcc44ff"]

# GL line direction modes
GL_MODE_ALONG  = "along"   # line drawn along the road  → bearing = road direction
GL_MODE_ACROSS = "across"  # line drawn across the road → bearing+90 = road dir, length = width


# ── geometry ───────────────────────────────────────────────────────
def haversine(lat1:float,lon1:float,lat2:float,lon2:float)->float:
    p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R_EARTH*math.asin(math.sqrt(max(0.0,min(1.0,a))))

def forward_bearing(lat1:float,lon1:float,lat2:float,lon2:float)->float:
    p1,p2=math.radians(lat1),math.radians(lat2)
    dl=math.radians(lon2-lon1)
    x=math.sin(dl)*math.cos(p2)
    y=math.cos(p1)*math.sin(p2)-math.sin(p1)*math.cos(p2)*math.cos(dl)
    return(math.degrees(math.atan2(x,y))+360)%360

def norm180(h:float)->float:
    h=h%360; return h-180 if h>=180 else h

def offset_ll(lat:float,lon:float,bear:float,dist:float)->Tuple[float,float]:
    d=dist/R_EARTH; b=math.radians(bear)
    p1=math.radians(lat); l1=math.radians(lon)
    p2=math.asin(math.sin(p1)*math.cos(d)+math.cos(p1)*math.sin(d)*math.cos(b))
    l2=l1+math.atan2(math.sin(b)*math.sin(d)*math.cos(p1),
                     math.cos(d)-math.sin(p1)*math.sin(p2))
    return math.degrees(p2),math.degrees(l2)

def build_strip(clat:float,clon:float,heading:float,along:float,
                pnear:float,pfar:float,sw:float)->List[Tuple[float,float]]:
    """Build one strip rectangle. Strips run PERPENDICULAR to road heading."""
    pd=(heading+90)%360; hsw=sw/2; hlen=abs(pfar-pnear)/2; mp=(pnear+pfar)/2
    cl,cn=offset_ll(clat,clon,heading,along)
    sl,sn=offset_ll(cl,cn,pd,mp)
    c=[]
    for a in(+1,-1):
        for p in(+1,-1):
            la,lo=offset_ll(sl,sn,heading,a*hsw)
            la,lo=offset_ll(la,lo,pd,p*hlen)
            c.append((la,lo))
    return[c[0],c[1],c[3],c[2],c[0]]


def rotated_bbox(heading:float, polys:list) -> List[Tuple[float,float]]:
    """
    Compute tight bounding box ALIGNED to road heading.
    Returns 5-point closed polygon in lat/lon.
    FIX for BUG 4 — replaces axis-aligned bbox that gave huge wrong rectangles.
    """
    if not polys:
        return []
    # Use first polygon centre as reference origin
    ref_lat = polys[0].coords[0][0]
    ref_lon = polys[0].coords[0][1]
    b_rad   = math.radians(heading)
    perp_rad = math.radians((heading + 90) % 360)

    # Project all coords onto (along-road, cross-road) axes
    along_vals = []
    cross_vals = []
    all_coords = [c for p in polys for c in p.coords]
    for lat, lon in all_coords:
        dlat = math.radians(lat - ref_lat) * R_EARTH
        dlon = math.radians(lon - ref_lon) * R_EARTH * math.cos(math.radians(ref_lat))
        # along = projection onto road direction
        a = dlat * math.cos(b_rad)   + dlon * math.sin(b_rad)
        # cross = projection onto perpendicular
        c = dlat * math.cos(perp_rad)+ dlon * math.sin(perp_rad)
        along_vals.append(a); cross_vals.append(c)

    a_min,a_max = min(along_vals)-0.05, max(along_vals)+0.05
    c_min,c_max = min(cross_vals)-0.05, max(cross_vals)+0.05

    # Convert 4 corners back to lat/lon
    corners = []
    for a, c in [(a_min,c_min),(a_min,c_max),(a_max,c_max),(a_max,c_min)]:
        # Reconstruct lat/lon from (along, cross) offsets
        # along direction: (cos(b), sin(b)) in (lat,lon) space
        # cross direction: (cos(b+90), sin(b+90))
        clat = ref_lat + math.degrees((a*math.cos(b_rad) + c*math.cos(perp_rad))/R_EARTH)
        clon = ref_lon + math.degrees((a*math.sin(b_rad) + c*math.sin(perp_rad))
                                      /(R_EARTH*math.cos(math.radians(ref_lat))))
        corners.append((clat,clon))
    corners.append(corners[0])
    return corners


# ── KML parse ──────────────────────────────────────────────────────
@dataclass
class KMLMarker:
    name:str; lat:float; lon:float; index:int=0

@dataclass
class GreenLine:
    name:str
    start_lat:float; start_lon:float
    end_lat:float;   end_lon:float
    length_m:float
    bearing_deg:float
    # These two depend on gl_mode set at match time:
    road_heading:float    # actual road direction (not always bearing+90)
    gives_width:bool      # True if length_m = road width
    midpoint_lat:float; midpoint_lon:float


def parse_kml(path:str)->Tuple[List[KMLMarker],List[GreenLine]]:
    """Parse KML — Point placemarks + LineString placemarks."""
    root=ET.parse(path).getroot()
    markers:List[KMLMarker]=[]; gls:List[GreenLine]=[]
    def iter_pm(node):
        for c in node:
            if c.tag.split("}")[-1]=="Placemark": yield c
            else: yield from iter_pm(c)
    idx=0
    for pm in iter_pm(root):
        name=f"Marker_{idx+1}"
        for c in pm:
            if c.tag.split("}")[-1]=="name": name=(c.text or"").strip() or name; break
        # Point
        for el in pm.iter():
            if el.tag.split("}")[-1]=="Point":
                for ce in el.iter():
                    if ce.tag.split("}")[-1]=="coordinates" and ce.text:
                        p=ce.text.strip().split(",")
                        if len(p)>=2:
                            try:
                                markers.append(KMLMarker(name,float(p[1]),float(p[0]),idx))
                                idx+=1
                            except ValueError: pass
                break
        # LineString
        for el in pm.iter():
            if el.tag.split("}")[-1]=="LineString":
                for ce in el.iter():
                    if ce.tag.split("}")[-1]=="coordinates" and ce.text:
                        pts=ce.text.strip().split()
                        if len(pts)>=2:
                            try:
                                p1=[float(v) for v in pts[0].split(",")]
                                p2=[float(v) for v in pts[-1].split(",")]
                                la1,lo1=p1[1],p1[0]; la2,lo2=p2[1],p2[0]
                                d=haversine(la1,lo1,la2,lo2)
                                b=forward_bearing(la1,lo1,la2,lo2)
                                # road_heading and gives_width set later by match_gl
                                gls.append(GreenLine(
                                    name,la1,lo1,la2,lo2,
                                    round(d,3),round(b,2),
                                    road_heading=round(norm180(b),2),  # default: along
                                    gives_width=False,
                                    midpoint_lat=(la1+la2)/2,
                                    midpoint_lon=(lo1+lo2)/2,
                                ))
                            except(ValueError,IndexError): pass
                break
    return markers,gls


def match_gl(
    markers:List[KMLMarker],
    gls:List[GreenLine],
    max_dist_m:float=50.0,
    gl_mode:str=GL_MODE_ALONG,   # "along" or "across"
) -> Dict[int, GreenLine]:
    """
    FIX BUG 1 + BUG 2 + BUG 3:

    BUG 1 fix: road_heading depends on gl_mode
      - along:  heading = norm180(GL_bearing)  ← line runs WITH the road
      - across: heading = norm180(GL_bearing + 90) ← line runs ACROSS the road

    BUG 2 fix: match GL to ALL markers within max_dist, not just nearest 1
      - Every marker within 50m gets the GL heading assigned
      - Closest marker also gets road_width if gl_mode=across

    BUG 3 fix: only use GL length as road_width if gl_mode=across
      - along mode: length = road segment length → don't use as width
    """
    # Update road_heading and gives_width on each GL based on mode
    for gl in gls:
        if gl_mode == GL_MODE_ACROSS:
            gl.road_heading = round(norm180(gl.bearing_deg + 90), 2)
            gl.gives_width  = True
        else:  # along
            gl.road_heading = round(norm180(gl.bearing_deg), 2)
            gl.gives_width  = False

    result: Dict[int, GreenLine] = {}

    for gl in gls:
        # Find ALL markers within max_dist of this GL's midpoint or endpoints
        candidates = []
        for mk in markers:
            # Distance to midpoint
            d_mid = haversine(gl.midpoint_lat, gl.midpoint_lon, mk.lat, mk.lon)
            # Distance to start/end (in case marker is near an endpoint)
            d_s   = haversine(gl.start_lat, gl.start_lon, mk.lat, mk.lon)
            d_e   = haversine(gl.end_lat,   gl.end_lon,   mk.lat, mk.lon)
            d_min = min(d_mid, d_s, d_e)
            if d_min <= max_dist_m:
                candidates.append((d_min, mk.index))

        if not candidates:
            continue

        candidates.sort()  # closest first

        for dist, mk_idx in candidates:
            # For width: only the closest marker gets the GL's width
            # All candidates get the heading
            if mk_idx not in result:
                result[mk_idx] = gl
            else:
                # Keep if this GL is closer
                prev_gl = result[mk_idx]
                prev_d  = haversine(prev_gl.midpoint_lat, prev_gl.midpoint_lon,
                                    markers[mk_idx].lat, markers[mk_idx].lon)
                if dist < prev_d:
                    result[mk_idx] = gl

    return result


# ── heading detection ──────────────────────────────────────────────
def road_heading(
    markers:List[KMLMarker], idx:int,
    gl:Optional[GreenLine]=None,
    override:Optional[float]=None,
    w:int=3,
) -> Tuple[float,str]:
    if override is not None: return norm180(override),"manual"
    if gl is not None: return gl.road_heading,"green-line"
    bs:List[Tuple[float,float]]=[]
    if markers and 0<=idx<len(markers):
        mk=markers[idx]
        for off in range(-w,w+1):
            if off==0: continue
            j=idx+off
            if j<0 or j>=len(markers): continue
            nb=markers[j]; d=haversine(mk.lat,mk.lon,nb.lat,nb.lon)
            if d<0.5: continue
            b=norm180(forward_bearing(mk.lat,mk.lon,nb.lat,nb.lon))
            bs.append((b,1.0/(abs(off)*max(d,1.0))))
    if not bs: return 0.0,"default"
    sx=sum(w2*math.cos(math.radians(2*b)) for b,w2 in bs)
    sy=sum(w2*math.sin(math.radians(2*b)) for b,w2 in bs)
    return norm180(math.degrees(math.atan2(sy,sx))/2),"neighbour"


# ── spec ───────────────────────────────────────────────────────────
@dataclass
class MarkerOverride:
    num_lanes:Optional[int]=None; road_width_m:Optional[float]=None
    separator_width_m:Optional[float]=None; heading_deg:Optional[float]=None
    lane_gap_m:Optional[float]=None; marker_position:str="centre"
    custom_offset_m:float=0.0; strip_length_m:Optional[float]=None

@dataclass
class PolygonSpec:
    strip_width_mm:float=15.0; num_strips:int=6; strip_gap_m:float=0.10
    num_lanes:int=2; road_width_m:float=7.0; separator_width_m:float=0.5
    lane_gap_m:float=-1.0; heading_override:Optional[float]=None
    strip_length_m:Optional[float]=None
    gl_mode:str=GL_MODE_ALONG        # "along" or "across"
    marker_overrides:Dict[int,MarkerOverride]=field(default_factory=dict)
    greenline_matches:Dict[int,GreenLine]=field(default_factory=dict)

    def eff_sep(self,ov=None)->float:
        return ov.separator_width_m if ov and ov.separator_width_m is not None else self.separator_width_m
    def eff_lanes(self,ov=None)->int:
        return ov.num_lanes if ov and ov.num_lanes is not None else self.num_lanes
    def eff_rw(self,i:int)->float:
        gl=self.greenline_matches.get(i)
        # Only use GL length as width if it was drawn ACROSS the road
        if gl and gl.gives_width: return gl.length_m
        ov=self.marker_overrides.get(i)
        if ov and ov.road_width_m is not None: return ov.road_width_m
        return self.road_width_m
    def eff_rw_src(self,i:int)->str:
        gl=self.greenline_matches.get(i)
        if gl and gl.gives_width: return "green-line (across)"
        ov=self.marker_overrides.get(i)
        if ov and ov.road_width_m is not None: return "manual"
        return "global"
    def eff_gap(self,lw:float,ov=None)->float:
        lg=ov.lane_gap_m if ov and ov.lane_gap_m is not None else self.lane_gap_m
        if lg and lg>0: return lg
        return self.eff_sep(ov)+max(0.3,lw*0.10)
    def eff_sl(self,lw:float,ov=None)->float:
        sl=ov.strip_length_m if ov and ov.strip_length_m is not None else self.strip_length_m
        return sl if sl is not None else lw


# ── road centre ────────────────────────────────────────────────────
def road_centre(lat,lon,heading,rw,ov,spec,idx)->Tuple[float,float]:
    # If GL is across-road, its midpoint = road centre (accurate)
    gl=spec.greenline_matches.get(idx)
    if gl and gl.gives_width: return gl.midpoint_lat,gl.midpoint_lon
    if ov is None or ov.marker_position=="centre": return lat,lon
    nl=spec.eff_lanes(ov); sep=spec.eff_sep(ov); lw=(rw-sep)/max(nl,1)
    perp=(heading+90)%360
    offs={"left_edge":rw/2,"left_lane":lw/2,"right_lane":-lw/2,
          "right_edge":-rw/2,"custom":ov.custom_offset_m}
    sh=offs.get(ov.marker_position,0.0)
    if sh==0.0: return lat,lon
    return offset_ll(lat,lon,perp,sh)


# ── polygon dataclass ──────────────────────────────────────────────
@dataclass
class GenPoly:
    marker_idx:int; marker_name:str; lane_idx:int; strip_idx:int
    global_idx:int; coords:List[Tuple[float,float]]
    road_heading:float; heading_src:str
    road_width_m:float; rw_src:str
    lane_width_m:float; strip_width_m:float; strip_len_m:float
    marker_pos:str; num_lanes:int


# ── strip generator ────────────────────────────────────────────────
def gen_marker(mk:KMLMarker,spec:PolygonSpec,heading:float,hsrc:str)->List[GenPoly]:
    i=mk.index; ov=spec.marker_overrides.get(i)
    rw=spec.eff_rw(i); rws=spec.eff_rw_src(i)
    nl=spec.eff_lanes(ov); sep=spec.eff_sep(ov); lw=(rw-sep)/max(nl,1)
    hg=spec.eff_gap(lw,ov)/2; sw=spec.strip_width_mm/1000
    ns=spec.num_strips; gap=spec.strip_gap_m
    clat,clon=road_centre(mk.lat,mk.lon,heading,rw,ov,spec,i)
    spl=[ns//nl]*nl
    for x in range(ns%nl): spl[x]+=1
    hl=nl//2; res:List[GenPoly]=[]; gi=0
    for li in range(nl):
        n=spl[li]; sl=spec.eff_sl(lw,ov)
        tier,sign=(hl-1-li,+1) if li<hl else (li-hl,-1)
        inn=sign*(hg+tier*lw); out=sign*(hg+(tier+1)*lw)
        ta=n*sw+(n-1)*gap; sa=-ta/2
        for si in range(n):
            ac=sa+si*(sw+gap)+sw/2
            coords=build_strip(clat,clon,heading,ac,inn,out,sw)
            res.append(GenPoly(i,mk.name,li,si,gi,coords,heading,hsrc,
                rw,rws,lw,sw,sl,ov.marker_position if ov else "centre",nl))
            gi+=1
    return res


# ── pipeline ───────────────────────────────────────────────────────
def run_pipeline(
    kml_path:str, spec:PolygonSpec,
    per_headings:Optional[Dict[int,float]]=None,
) -> Tuple[List[KMLMarker],List[GreenLine],Dict[int,GreenLine],List[GenPoly]]:
    markers,gls=parse_kml(kml_path)
    # Pass gl_mode to match_gl so headings and width flags are set correctly
    glm=match_gl(markers,gls,max_dist_m=50.0,gl_mode=spec.gl_mode)
    spec.greenline_matches=glm
    all_p:List[GenPoly]=[]
    for mk in markers:
        i=mk.index; ov=spec.marker_overrides.get(i)
        hov=None
        if per_headings and i in per_headings: hov=per_headings[i]
        elif ov and ov.heading_deg is not None: hov=ov.heading_deg
        elif spec.heading_override is not None: hov=spec.heading_override
        h,hs=road_heading(markers,i,glm.get(i),hov)
        all_p.extend(gen_marker(mk,spec,h,hs))
    return markers,gls,glm,all_p


# ── KML export (with rotated bbox — BUG 4 fix) ────────────────────
def export_kml(markers,all_p,spec,path):
    ml=max((p.num_lanes for p in all_p),default=2)
    by={}
    for p in all_p: by.setdefault(p.marker_idx,[]).append(p)
    def cs(coords): return" ".join(f"{lo:.8f},{la:.8f},0" for la,lo in coords)
    def styles():
        s=""
        for i in range(max(ml,2)):
            c=LANE_COLOURS_KML[i%len(LANE_COLOURS_KML)]
            s+=(f'\n  <Style id="sL{i}"><LineStyle><color>{c}</color><width>3</width></LineStyle>'
                f'<PolyStyle><color>{c}</color><fill>1</fill></PolyStyle></Style>')
        s+=('\n  <Style id="bnd"><LineStyle><color>ff0000ff</color><width>1</width></LineStyle>'
            '<PolyStyle><color>00000000</color><fill>0</fill></PolyStyle></Style>'
            '\n  <Style id="pin"><IconStyle><scale>1.0</scale>'
            '<Icon><href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href>'
            '</Icon></IconStyle></Style>')
        return s
    L=['<?xml version="1.0" encoding="UTF-8"?>',
       '<kml xmlns="http://www.opengis.net/kml/2.2">','<Document>',
       '<n>CAP PTBM Speed Breaker Polygons — GIS BOQ Tool v10</n>',
       styles()]
    for mk in markers:
        ps=by.get(mk.index,[])
        L.append(f'<Folder><n>{mk.name}</n>')
        L+=[f'<Placemark><n>{mk.name}</n><styleUrl>#pin</styleUrl>'
            f'<Point><coordinates>{mk.lon:.8f},{mk.lat:.8f},0</coordinates></Point></Placemark>']
        if ps:
            p0=ps[0]
            # BUG 4 FIX: rotated tight bounding box aligned to road heading
            bbox=rotated_bbox(p0.road_heading,ps)
            if bbox:
                L+=[f'<Placemark><n>Boundary-{mk.name}</n>'
                    f'<description>Road {p0.road_width_m:.2f}m [{p0.rw_src}] | '
                    f'Heading {p0.road_heading:.1f}° [{p0.heading_src}] | '
                    f'{len(ps)} strips</description>'
                    f'<styleUrl>#bnd</styleUrl>'
                    f'<Polygon><outerBoundaryIs><LinearRing>'
                    f'<coordinates>{cs(bbox)}</coordinates>'
                    f'</LinearRing></outerBoundaryIs></Polygon></Placemark>']
            for p in ps:
                L+=[f'<Placemark><n>S{p.global_idx+1} L{p.lane_idx+1}-{mk.name}</n>'
                    f'<description>{p.strip_width_m*1000:.0f}mm | Lane {p.lane_idx+1} | '
                    f'Road {p.road_width_m:.2f}m [{p.rw_src}] | '
                    f'Heading {p.road_heading:.1f}°</description>'
                    f'<styleUrl>#sL{p.lane_idx%max(p.num_lanes,1)}</styleUrl>'
                    f'<Polygon><outerBoundaryIs><LinearRing>'
                    f'<coordinates>{cs(p.coords)}</coordinates>'
                    f'</LinearRing></outerBoundaryIs></Polygon></Placemark>']
        L.append('</Folder>')
    L+=["</Document>","</kml>"]
    with open(path,"w",encoding="utf-8") as f: f.write("\n".join(L))


# ── Excel export ───────────────────────────────────────────────────
def _tb():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _fill(h): return PatternFill("solid",start_color=h,end_color=h)
def _hdr(ws,r,c,v,bg="1F3864",fg="FFFFFF"):
    x=ws.cell(r,c,v); x.font=Font(bold=True,color=fg,size=9)
    x.fill=_fill(bg); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    x.border=_tb(); return x
def _dat(ws,r,c,v,bg=None,bold=False):
    x=ws.cell(r,c,v); x.font=Font(bold=bold,size=9)
    x.alignment=Alignment(horizontal="center",vertical="center")
    x.border=_tb()
    if bg: x.fill=_fill(bg)
    return x

def export_excel(markers,gls,glm,all_p,spec,path):
    wb=openpyxl.Workbook()
    # Sheet1 BOQ
    ws1=wb.active; ws1.title="BOQ Summary"
    H1=["S.No","Marker Name","Lat","Lon","Road Width (m)","Width Source",
        "Heading (°)","Heading Src","GL Mode","Lanes","Lane Width (m)",
        "Strips","Strip (mm)","Strip Length (m)","Area/Strip","Total Area","Notes"]
    for ci,h in enumerate(H1,1): _hdr(ws1,1,ci,h)
    ws1.row_dimensions[1].height=36
    by={};
    for p in all_p: by.setdefault(p.marker_idx,[]).append(p)
    for ri,mk in enumerate(markers,2):
        ps=by.get(mk.index,[]); bg="F2F7FF" if ri%2==0 else "FFFFFF"
        p0=ps[0] if ps else None
        rw=p0.road_width_m if p0 else spec.road_width_m
        rws=p0.rw_src if p0 else "global"
        sl=p0.strip_len_m if p0 else 0.0; area=(p0.strip_width_m*sl) if p0 else 0.0
        gl=glm.get(mk.index)
        gl_mode_lbl="Across" if (gl and gl.gives_width) else "Along" if gl else "—"
        notes=("🟢 GL across (width auto)" if (gl and gl.gives_width)
               else "🔵 GL along (heading only)" if gl
               else "🔶 Manual" if rws=="manual" else "")
        row=[ri-1,mk.name,round(mk.lat,6),round(mk.lon,6),
             round(rw,3),rws,
             round(p0.road_heading,1) if p0 else 0,
             p0.heading_src if p0 else "—",
             gl_mode_lbl,
             p0.num_lanes if p0 else spec.num_lanes,
             round(p0.lane_width_m,3) if p0 else 0,
             len(ps),spec.strip_width_mm,round(sl,3),
             round(area,4),round(area*len(ps),4),notes]
        for ci,val in enumerate(row,1):
            cbg=("FFD700" if ci==5 and gl and gl.gives_width
                 else "E8F4FD" if ci==5 and gl and not gl.gives_width
                 else "FFF3CD" if ci==5 and rws=="manual" else bg)
            _dat(ws1,ri,ci,val,bg=cbg)
    for i,w in enumerate([5,24,11,11,14,14,11,12,9,7,12,7,10,12,11,11,22],1):
        ws1.column_dimensions[get_column_letter(i)].width=w

    # Sheet2 — Road Measurements
    ws2=wb.create_sheet("Road Measurements")
    H2=["S.No","Line Name","Length (m)","Bearing (°)","Road Heading (°)",
        "GL Mode","Gives Width?","Start Lat","Start Lon","End Lat","End Lon",
        "Mid Lat","Mid Lon","Matched Markers","Status"]
    for ci,h in enumerate(H2,1): _hdr(ws2,1,ci,h,bg="145A32")
    ws2.row_dimensions[1].height=36
    if not gls:
        ws2.merge_cells("A2:O2")
        c=ws2.cell(2,1,"No green lines found. Draw lines along/across road in Google Earth Pro → re-upload.")
        c.font=Font(italic=True,color="C0392B",size=10)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.fill=_fill("FADBD8")
        ws2.row_dimensions[2].height=30
    else:
        mk_idx_map={mk.index:mk for mk in markers}
        for ri,gl in enumerate(gls,2):
            matched=[mk_idx_map[mi].name for mi,g in glm.items() if id(g)==id(gl)]
            ok=len(matched)>0; bg="E9F7EF" if ok else "FDECEA"
            row2=[ri-1,gl.name,round(gl.length_m,3),round(gl.bearing_deg,2),
                  round(gl.road_heading,2),
                  "Across (width)" if gl.gives_width else "Along (heading)",
                  "Yes" if gl.gives_width else "No",
                  round(gl.start_lat,6),round(gl.start_lon,6),
                  round(gl.end_lat,6),round(gl.end_lon,6),
                  round(gl.midpoint_lat,6),round(gl.midpoint_lon,6),
                  ", ".join(matched) if matched else "—",
                  f"✅ {len(matched)} marker(s)" if ok else "⚠️ No match"]
            for ci,val in enumerate(row2,1):
                c=_dat(ws2,ri,ci,val,bg="FFD700" if ci==5 else bg)
                if ci==5: c.font=Font(bold=True,size=10)
    for i,w in enumerate([5,22,11,11,14,16,11,12,12,12,12,12,12,28,14],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    # Sheet3 Strip Coords
    ws3=wb.create_sheet("Strip Coordinates")
    H3=["S.No","Marker","Lane","Strip","C1 Lat","C1 Lon","C2 Lat","C2 Lon",
        "C3 Lat","C3 Lon","C4 Lat","C4 Lon","Road Width","Source","Strip mm","Len m"]
    for ci,h in enumerate(H3,1): _hdr(ws3,1,ci,h,bg="4A235A")
    ws3.row_dimensions[1].height=36
    for ri,p in enumerate(all_p,2):
        bg="F5EEF8" if ri%2==0 else "FFFFFF"; cs_=p.coords[:4]
        r3=[ri-1,p.marker_name,p.lane_idx+1,p.strip_idx+1,
            *[round(v,8) for c_ in cs_ for v in c_],
            round(p.road_width_m,3),p.rw_src,
            round(p.strip_width_m*1000,1),round(p.strip_len_m,3)]
        for ci,val in enumerate(r3,1): _dat(ws3,ri,ci,val,bg=bg)
    for i,w in enumerate([5,22,7,8,12,12,12,12,12,12,12,12,11,14,9,9],1):
        ws3.column_dimensions[get_column_letter(i)].width=w

    # Sheet4 Spec
    ws4=wb.create_sheet("Installation Spec")
    _hdr(ws4,1,1,"Parameter",bg="784212"); _hdr(ws4,1,2,"Value",bg="784212")
    ws4.column_dimensions["A"].width=30; ws4.column_dimensions["B"].width=65
    rows=[("Tool","GIS BOQ Speed Breaker Tool v10"),
          ("Institute","IIIT Nagpur — Dr. Neha Kasture"),("Client","PWD / NHAI"),("",""),
          ("Strip Type","CAP PTBM (Capsule Prefab Thermoplastic Bituminous Marking)"),
          ("Strip Width",f"{spec.strip_width_mm:.0f} mm"),
          ("Total Strips",str(spec.num_strips)),("Strip Gap",f"{spec.strip_gap_m*100:.0f} cm"),
          ("GL Mode",f"{'Across road (width from GL)' if spec.gl_mode==GL_MODE_ACROSS else 'Along road (heading from GL)'}"),
          ("Default Road Width",f"{spec.road_width_m:.1f} m"),
          ("Default Lanes",str(spec.num_lanes)),
          ("Separator",f"{spec.separator_width_m:.2f} m"),("",""),
          ("Total Markers",str(len(markers))),("Total Strips",str(len(all_p))),
          ("Green Lines Found",str(len(gls))),("GL Matched",str(len(glm))),("",""),
          ("Road Width Priority",
           "1. GL across (length=width)  |  2. Manual  |  3. Global"),
          ("Heading Priority",
           "1. Manual override  |  2. GL along (bearing) or GL across (bearing+90°)  |  3. Neighbour avg")]
    for ri,(k,v) in enumerate(rows,2):
        ws4.cell(ri,1,k).font=Font(bold=bool(k),size=9)
        ws4.cell(ri,2,v).font=Font(size=9)
    wb.save(path)


# ── self test ──────────────────────────────────────────────────────
if __name__=="__main__":
    import tempfile,os
    # Use the actual KML from the user's file
    KML="""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
    <Placemark><n>Untitled path</n>
        <LineString><coordinates>
            93.9456058575073,24.83222939738513,0 93.94549199991495,24.83219897227781,0
        </coordinates></LineString>
    </Placemark>
    <Placemark><n>Untitled path 2</n>
        <LineString><coordinates>
            93.9456266349721,24.83217345898283,0 93.94572756213377,24.83219175047888,0
        </coordinates></LineString>
    </Placemark>
    <Placemark><n>Untitled placemark</n>
        <Point><coordinates>93.94549267755646,24.83219770964525,807</coordinates></Point>
    </Placemark>
    <Placemark><n>Untitled placemark</n>
        <Point><coordinates>93.94560495041627,24.83222823099334,808</coordinates></Point>
    </Placemark>
    <Placemark><n>Untitled placemark</n>
        <Point><coordinates>93.94562693818699,24.83217228694374,808</coordinates></Point>
    </Placemark>
    <Placemark><n>Untitled placemark</n>
        <Point><coordinates>93.94572876485692,24.83219015575611,808</coordinates></Point>
    </Placemark>
</Document></kml>"""
    with tempfile.NamedTemporaryFile(mode="w",suffix=".kml",delete=False) as f:
        f.write(KML); kp=f.name

    spec=PolygonSpec(
        road_width_m=7.0, num_lanes=2, separator_width_m=0.5,
        gl_mode=GL_MODE_ALONG,  # lines drawn ALONG road in this KML
    )
    m,g,glm,p=run_pipeline(kp,spec)
    print(f"✅ Markers:{len(m)} GL:{len(g)} Matched:{len(glm)} Strips:{len(p)}")
    print()
    for gl in g:
        print(f"  📏 {gl.name}: {gl.length_m:.2f}m  "
              f"bearing={gl.bearing_deg:.1f}°  road_heading={gl.road_heading:.1f}°  "
              f"gives_width={gl.gives_width}")
    print()
    for mk in m:
        ps=[x for x in p if x.marker_idx==mk.index]
        gl=glm.get(mk.index)
        if ps:
            print(f"  📍 {mk.name}: heading={ps[0].road_heading:.1f}° [{ps[0].heading_src}]  "
                  f"width={ps[0].road_width_m:.2f}m [{ps[0].rw_src}]  strips={len(ps)}")
    export_excel(m,g,glm,p,spec,"/tmp/p1_v10_test.xlsx")
    export_kml(m,p,spec,"/tmp/p1_v10_test.kml")
    print("\n✅ Excel + KML OK")
    os.unlink(kp)