"""
p2.py v3 — Speed Breaker CAP PTBM Polygon Engine
IIIT Nagpur | Under Dr. Neha Kasture | PWD / NHAI

FIXES in v3:
  ● Added lane_width_m to GenPoly (was causing AttributeError)
  ● Marker name label shown in KML (matches screenshot)
  ● Strips PERPENDICULAR to road spanning full road width
  ● Default gap = 0.5m (clearly visible in satellite)
"""
from __future__ import annotations
import math, xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

R_EARTH = 6_371_000.0

LANE_PRESETS: Dict[str, dict] = {
    "1-Lane (3.5m)":  dict(num_lanes=1, road_width_m=3.5,  separator_width_m=0.0),
    "2-Lane (7.0m)":  dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
    "4-Lane (14.0m)": dict(num_lanes=4, road_width_m=14.0, separator_width_m=2.0),
    "6-Lane (21.0m)": dict(num_lanes=6, road_width_m=21.0, separator_width_m=3.0),
    "Custom":         dict(num_lanes=2, road_width_m=7.0,  separator_width_m=0.5),
}
LANE_COLOURS_KML = ["ff00d7ff","ff0088ff","ff00ff88","ffff44aa","ff44ffcc","ffcc44ff"]

# ── geometry ───────────────────────────────────────────────────────
def haversine(lat1,lon1,lat2,lon2)->float:
    p1,p2=math.radians(lat1),math.radians(lat2)
    dp=math.radians(lat2-lat1); dl=math.radians(lon2-lon1)
    a=math.sin(dp/2)**2+math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2*R_EARTH*math.asin(math.sqrt(max(0.0,min(1.0,a))))

def forward_bearing(lat1,lon1,lat2,lon2)->float:
    p1,p2=math.radians(lat1),math.radians(lat2)
    dl=math.radians(lon2-lon1)
    x=math.sin(dl)*math.cos(p2)
    y=math.cos(p1)*math.sin(p2)-math.sin(p1)*math.cos(p2)*math.cos(dl)
    return(math.degrees(math.atan2(x,y))+360)%360

def norm180(h:float)->float:
    h=h%360; return h-180 if h>=180 else h

def offset_ll(lat,lon,bearing_deg,dist_m)->Tuple[float,float]:
    d=dist_m/R_EARTH; b=math.radians(bearing_deg)
    p1=math.radians(lat); l1=math.radians(lon)
    p2=math.asin(math.sin(p1)*math.cos(d)+math.cos(p1)*math.sin(d)*math.cos(b))
    l2=l1+math.atan2(math.sin(b)*math.sin(d)*math.cos(p1),math.cos(d)-math.sin(p1)*math.sin(p2))
    return math.degrees(p2),math.degrees(l2)

def build_perp_strip(centre_lat,centre_lon,road_heading,along_offset,road_width,strip_thick)->List[Tuple[float,float]]:
    sc_lat,sc_lon=offset_ll(centre_lat,centre_lon,road_heading,along_offset)
    perp=(road_heading+90)%360
    half_w=road_width/2; half_t=strip_thick/2
    left_lat,left_lon=offset_ll(sc_lat,sc_lon,perp,+half_w)
    right_lat,right_lon=offset_ll(sc_lat,sc_lon,perp,-half_w)
    lf=offset_ll(left_lat,left_lon,road_heading,+half_t)
    lb=offset_ll(left_lat,left_lon,road_heading,-half_t)
    rb=offset_ll(right_lat,right_lon,road_heading,-half_t)
    rf=offset_ll(right_lat,right_lon,road_heading,+half_t)
    return [lf,lb,rb,rf,lf]

# ── KML parse ──────────────────────────────────────────────────────
@dataclass
class KMLMarker:
    name:str; lat:float; lon:float; index:int=0

def parse_kml(path:str)->List[KMLMarker]:
    root=ET.parse(path).getroot(); markers=[]
    def iter_pm(node):
        for c in node:
            if c.tag.split("}")[-1]=="Placemark": yield c
            else: yield from iter_pm(c)
    idx=0
    for pm in iter_pm(root):
        name=f"Marker_{idx+1}"
        for c in pm:
            if c.tag.split("}")[-1]=="name": name=(c.text or"").strip() or name; break
        for el in pm.iter():
            if el.tag.split("}")[-1]=="Point":
                for ce in el.iter():
                    if ce.tag.split("}")[-1]=="coordinates" and ce.text:
                        p=ce.text.strip().split(",")
                        if len(p)>=2:
                            try: markers.append(KMLMarker(name,float(p[1]),float(p[0]),idx)); idx+=1
                            except ValueError: pass
                break
    return markers

# ── heading ────────────────────────────────────────────────────────
def detect_heading(markers,idx,window=3)->Tuple[float,str]:
    if len(markers)<2: return 0.0,"default (0°)"
    mk=markers[idx]; bearings=[]
    for off in range(-window,window+1):
        if off==0: continue
        j=idx+off
        if j<0 or j>=len(markers): continue
        nb=markers[j]; d=haversine(mk.lat,mk.lon,nb.lat,nb.lon)
        if d<0.5: continue
        b=norm180(forward_bearing(mk.lat,mk.lon,nb.lat,nb.lon))
        bearings.append((b,1.0/(abs(off)*max(d,1.0))))
    if not bearings: return 0.0,"default (0°)"
    sx=sum(w*math.cos(math.radians(2*b)) for b,w in bearings)
    sy=sum(w*math.sin(math.radians(2*b)) for b,w in bearings)
    return norm180(math.degrees(math.atan2(sy,sx))/2),"auto-neighbour"

# ── spec ───────────────────────────────────────────────────────────
@dataclass
class PolySpec:
    road_width_m:float=7.0; num_lanes:int=2; separator_width_m:float=0.5
    num_strips:int=3; strip_width_mm:float=15.0; strip_gap_m:float=0.50
    heading_override:Optional[float]=None
    marker_overrides:Dict[int,dict]=field(default_factory=dict)

# ── result ─────────────────────────────────────────────────────────
@dataclass
class GenPoly:
    marker_idx:int; marker_name:str; strip_idx:int
    coords:List[Tuple[float,float]]
    road_heading:float; heading_src:str
    road_width_m:float; lane_width_m:float   # ← FIXED: added lane_width_m
    strip_width_m:float; strip_gap_m:float
    num_lanes:int; along_offset_m:float

# ── generator ──────────────────────────────────────────────────────
def generate_polygons(markers,spec)->Tuple[List[GenPoly],Dict[int,Tuple[float,str]]]:
    all_polys=[]; headings={}
    for mk in markers:
        i=mk.index; ov=spec.marker_overrides.get(i,{})
        rw=float(ov.get("road_width_m",spec.road_width_m))
        nl=int(ov.get("num_lanes",spec.num_lanes))
        sep=float(ov.get("separator_width_m",spec.separator_width_m))
        n_strips=int(ov.get("num_strips",spec.num_strips))
        sw_m=spec.strip_width_mm/1000.0; gap=spec.strip_gap_m
        lw=(rw-sep)/max(nl,1)
        manual_h=ov.get("heading_deg",spec.heading_override)
        if manual_h is not None: heading=norm180(float(manual_h)); hsrc="manual"
        else: heading,hsrc=detect_heading(markers,i)
        headings[i]=(heading,hsrc)
        total_span=n_strips*sw_m+(n_strips-1)*gap
        first_pos=-total_span/2+sw_m/2
        for si in range(n_strips):
            along=first_pos+si*(sw_m+gap)
            coords=build_perp_strip(mk.lat,mk.lon,heading,along,rw,sw_m)
            all_polys.append(GenPoly(
                marker_idx=i,marker_name=mk.name,strip_idx=si,
                coords=coords,road_heading=heading,heading_src=hsrc,
                road_width_m=rw,lane_width_m=lw,
                strip_width_m=sw_m,strip_gap_m=gap,
                num_lanes=nl,along_offset_m=along,
            ))
    return all_polys,headings

# ── KML export ─────────────────────────────────────────────────────
def export_kml(markers,all_polys,headings,spec,out_path):
    ml=max((p.num_lanes for p in all_polys),default=2)
    def cs(coords): return" ".join(f"{lo:.8f},{la:.8f},0" for la,lo in coords)
    def styles():
        s=""
        for i in range(max(ml,2)):
            c=LANE_COLOURS_KML[i%len(LANE_COLOURS_KML)]
            s+=(f'\n  <Style id="sL{i}"><LineStyle><color>{c}</color><width>1</width></LineStyle>'
                f'<PolyStyle><color>{c}</color><fill>1</fill></PolyStyle></Style>')
        # Pin with visible name label (matches screenshot)
        s+=('\n  <Style id="pin"><IconStyle><scale>1.1</scale>'
            '<Icon><href>http://maps.google.com/mapfiles/kml/pushpin/red-pushpin.png</href></Icon>'
            '</IconStyle><LabelStyle><scale>0.85</scale></LabelStyle></Style>')
        return s
    by={}
    for p in all_polys: by.setdefault(p.marker_idx,[]).append(p)
    L=['<?xml version="1.0" encoding="UTF-8"?>',
       '<kml xmlns="http://www.opengis.net/kml/2.2">','<Document>',
       '<n>CAP PTBM Speed Breaker Polygons — GIS BOQ v3</n>',styles()]
    for mk in markers:
        ps=by.get(mk.index,[]); h,_=headings.get(mk.index,(0.0,"default"))
        rw=ps[0].road_width_m if ps else spec.road_width_m
        nl=ps[0].num_lanes if ps else spec.num_lanes
        sw_mm=ps[0].strip_width_m*1000 if ps else spec.strip_width_mm
        L.append(f'<Folder><n>{mk.name}</n>')
        # Marker pin — name label visible like in screenshot
        L.append(f'<Placemark><n>{mk.name}</n>'
                 f'<description>CAP PTBM {sw_mm:.0f}MM X {len(ps)} | '
                 f'Road: {rw:.1f}m | {nl} Lane | Heading: {h:.1f}°</description>'
                 f'<styleUrl>#pin</styleUrl>'
                 f'<Point><coordinates>{mk.lon:.8f},{mk.lat:.8f},0</coordinates></Point>'
                 f'</Placemark>')
        # Strip polygons with CAP PTBM label
        for p in ps:
            label=f"CAP PTBM {p.strip_width_m*1000:.0f}MM X {len(ps)}"
            L.append(f'<Placemark><n>{label}</n>'
                     f'<description>{mk.name} | Strip {p.strip_idx+1}/{len(ps)} | '
                     f'Road {p.road_width_m:.2f}m | Heading {p.road_heading:.1f}° [{p.heading_src}]'
                     f'</description>'
                     f'<styleUrl>#sL{p.strip_idx%max(p.num_lanes,1)}</styleUrl>'
                     f'<Polygon><outerBoundaryIs><LinearRing>'
                     f'<coordinates>{cs(p.coords)}</coordinates>'
                     f'</LinearRing></outerBoundaryIs></Polygon></Placemark>')
        L.append('</Folder>')
    L+=["</Document>","</kml>"]
    with open(out_path,"w",encoding="utf-8") as f: f.write("\n".join(L))
    print(f"✅ KML saved: {out_path}")

# ── Excel export ───────────────────────────────────────────────────
def _tb():
    s=Side(style="thin"); return Border(left=s,right=s,top=s,bottom=s)
def _fill(h): return PatternFill("solid",start_color=h,end_color=h)
def _hdr(ws,r,c,v,bg="1F3864",fg="FFFFFF"):
    x=ws.cell(r,c,v); x.font=Font(bold=True,color=fg,size=9)
    x.fill=_fill(bg); x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    x.border=_tb(); return x
def _dat(ws,r,c,v,bg=None,bold=False):
    x=ws.cell(r,c,v); x.font=Font(bold=bold,size=9)
    x.alignment=Alignment(horizontal="center",vertical="center")
    x.border=_tb()
    if bg: x.fill=_fill(bg)
    return x

def export_excel(markers,all_polys,headings,spec,out_path):
    wb=openpyxl.Workbook()
    ws1=wb.active; ws1.title="BOQ Summary"
    H1=["S.No","Marker Name","Lat","Lon","Road Width (m)","Lanes","Lane Width (m)",
        "Heading (°)","Heading Source","Total Strips","Strip Thick (mm)","Gap (m)",
        "Area/Strip (m²)","Total Area (m²)","Total Span (m)"]
    for ci,h in enumerate(H1,1): _hdr(ws1,1,ci,h)
    ws1.row_dimensions[1].height=36
    by={}
    for p in all_polys: by.setdefault(p.marker_idx,[]).append(p)
    for ri,mk in enumerate(markers,2):
        ps=by.get(mk.index,[]); bg="F2F7FF" if ri%2==0 else "FFFFFF"
        p0=ps[0] if ps else None; h,hsrc=headings.get(mk.index,(0.0,"default"))
        rw=p0.road_width_m if p0 else spec.road_width_m
        nl=p0.num_lanes if p0 else spec.num_lanes
        lw=p0.lane_width_m if p0 else (rw-spec.separator_width_m)/max(nl,1)
        sw=p0.strip_width_m if p0 else spec.strip_width_mm/1000
        gap=p0.strip_gap_m if p0 else spec.strip_gap_m
        n=len(ps); area_s=sw*rw; area_t=area_s*n
        span=n*sw+max(n-1,0)*gap
        row=[ri-1,mk.name,round(mk.lat,6),round(mk.lon,6),round(rw,3),nl,round(lw,3),
             round(h,1),hsrc,n,round(sw*1000,1),round(gap,3),round(area_s,4),round(area_t,4),round(span,4)]
        for ci,val in enumerate(row,1): _dat(ws1,ri,ci,val,bg=bg)
    for i,w in enumerate([5,24,12,12,13,7,12,10,16,11,14,9,13,13,13],1):
        ws1.column_dimensions[get_column_letter(i)].width=w

    ws2=wb.create_sheet("Strip Coordinates")
    H2=["S.No","Marker","Strip No","Along Offset (m)",
        "C1 Lat","C1 Lon","C2 Lat","C2 Lon","C3 Lat","C3 Lon","C4 Lat","C4 Lon",
        "Road Width (m)","Strip mm","Gap (m)","Heading °"]
    for ci,h in enumerate(H2,1): _hdr(ws2,1,ci,h,bg="145A32")
    ws2.row_dimensions[1].height=36
    gi=1
    for p in all_polys:
        ri=gi+1; bg="E9F7EF" if ri%2==0 else "FFFFFF"; cs_=p.coords[:4]
        row2=[gi,p.marker_name,p.strip_idx+1,round(p.along_offset_m,4),
              *[round(v,8) for c_ in cs_ for v in c_],
              round(p.road_width_m,3),round(p.strip_width_m*1000,1),
              round(p.strip_gap_m,3),round(p.road_heading,1)]
        for ci,val in enumerate(row2,1): _dat(ws2,ri,ci,val,bg=bg)
        gi+=1
    for i,w in enumerate([5,22,9,14,12,12,12,12,12,12,12,12,12,9,9,10],1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    ws3=wb.create_sheet("Spec Summary")
    _hdr(ws3,1,1,"Parameter",bg="784212"); _hdr(ws3,1,2,"Value",bg="784212")
    ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=62
    rows=[("Tool","GIS BOQ Speed Breaker Tool v3 (p2)"),
          ("Institute","IIIT Nagpur — Dr. Neha Kasture"),("Client","PWD / NHAI"),("",""),
          ("Strip Type","CAP PTBM (Capsule Prefab Thermoplastic Bituminous Marking)"),
          ("Strip Orientation","PERPENDICULAR to road (spans full road width)"),
          ("Strip Thickness",f"{spec.strip_width_mm:.0f} mm  (along road direction)"),
          ("Strip Width",f"{spec.road_width_m:.1f} m  (cross-road = full road width)"),
          ("Number of Strips",str(spec.num_strips)),
          ("Gap Between Strips",f"{spec.strip_gap_m:.2f} m  (visible space along road)"),
          ("Default Road Width",f"{spec.road_width_m:.1f} m"),
          ("Default Lanes",str(spec.num_lanes)),("Separator Width",f"{spec.separator_width_m:.2f} m"),
          ("",""),("Total Markers",str(len(markers))),("Total Strips",str(len(all_polys))),
          ("",""),("IMPORTANT","All markers MUST be at CENTER of road. Strip spans ±(road_width/2) cross-road.")]
    for ri,(k,v) in enumerate(rows,2):
        ws3.cell(ri,1,k).font=Font(bold=bool(k),size=9)
        ws3.cell(ri,2,v).font=Font(size=9)
    wb.save(out_path)
    print(f"✅ Excel saved: {out_path}")

# ── pipeline ───────────────────────────────────────────────────────
def run_pipeline(kml_path,spec,per_headings=None):
    markers=parse_kml(kml_path)
    if not markers: raise ValueError("No Point markers found in KML.")
    if per_headings:
        for idx,hdg in per_headings.items():
            spec.marker_overrides.setdefault(idx,{})["heading_deg"]=hdg
    all_polys,headings=generate_polygons(markers,spec)
    return markers,all_polys,headings

# ── self test ──────────────────────────────────────────────────────
if __name__=="__main__":
    import tempfile,os
    TEST_KML="""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2"><Document>
  <Placemark><n>SB_1</n><Point><coordinates>93.94313688,24.83678499,0</coordinates></Point></Placemark>
  <Placemark><n>SB_2</n><Point><coordinates>93.94334266,24.83687056,0</coordinates></Point></Placemark>
  <Placemark><n>SB_3</n><Point><coordinates>93.94342449,24.83672434,0</coordinates></Point></Placemark>
</Document></kml>"""
    with tempfile.NamedTemporaryFile(mode="w",suffix=".kml",delete=False) as f:
        f.write(TEST_KML); kp=f.name
    spec=PolySpec(road_width_m=7.0,num_lanes=2,separator_width_m=0.5,
                  num_strips=3,strip_width_mm=15.0,strip_gap_m=0.50)
    markers,polys,headings=run_pipeline(kp,spec)
    print(f"\n✅ Markers:{len(markers)} | Total strips:{len(polys)}\n")
    for mk in markers:
        ps=[p for p in polys if p.marker_idx==mk.index]
        h,src=headings[mk.index]; p0=ps[0]
        print(f"  📍 {mk.name}  lat={mk.lat:.6f}  lon={mk.lon:.6f}")
        print(f"     heading={h:.1f}° [{src}]  road_width={p0.road_width_m:.1f}m  lane_width={p0.lane_width_m:.2f}m")
        for p in ps: print(f"       Strip {p.strip_idx+1}: offset={p.along_offset_m:+.4f}m")
        print()
    export_kml(markers,polys,headings,spec,"/tmp/p2v3_test.kml")
    export_excel(markers,polys,headings,spec,"/tmp/p2v3_test.xlsx")
    os.unlink(kp)
    print("✅ Done!")